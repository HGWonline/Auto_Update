import os
import time
import requests
from dotenv import load_dotenv
from datetime import datetime
import openpyxl
import re
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup

# -----------------------
# 1. 환경 변수 로드
# -----------------------
load_dotenv()

# Shopify API 인증 정보
SHOPIFY_ADMIN_TOKEN = os.getenv('SHOPIFY_ADMIN_TOKEN')
SHOPIFY_STORE_URL = os.getenv('SHOPIFY_STORE_URL')

# Shopify API 헤더 설정
headers = {
    "Content-Type": "application/json",
    "X-Shopify-Access-Token": SHOPIFY_ADMIN_TOKEN
}

# -----------------------
# 2. 스크래핑 코드
# -----------------------

def wait_for_page_load(driver, timeout=10):
    """ 페이지 로드가 완료될 때까지 대기하는 함수 """
    WebDriverWait(driver, timeout).until(
        lambda d: d.execute_script('return document.readyState') == 'complete'
    )

# 로그인 및 네비게이션
def login_and_navigate(driver):
    print("Navigating to login page...")
    driver.maximize_window()  # 브라우저 전체화면으로 설정
    driver.get('https://www.hangawee.com.au/login')

    # 페이지가 완전히 로드될 때까지 대기
    wait_for_page_load(driver)

    print("Entering login code...")
    code_input = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, 'username'))
    )
    code_input.send_keys('245554073198')

    # Sign in 버튼 클릭
    sign_in_button = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.ID, 'kt_login_signin_submit'))
    )
    print("Clicking sign in button...")
    sign_in_button.click()

    # Next 버튼 클릭
    next_button = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.ID, 'next-step'))
    )
    print("Clicking Next button...")
    next_button.click()

    # 팝업에서 Confirm 버튼 클릭
    confirm_button = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.ID, 'goIndex'))
    )
    print("Clicking Confirm button on popup...")
    confirm_button.click()

    # 홈 화면으로 이동한 후, itemlist 페이지로 이동
    print("Navigating to item list page...")
    driver.get('https://www.hangawee.com.au/?page=mitems')

    # 드롭다운이 로드되었는지 확인
    WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, 'button.btn.dropdown-toggle.btn-light'))
    )

    # JavaScript로 항목 수를 700개로 설정
    print("Setting item count to 700 using JavaScript...")
    dropdown_script = '''
    document.querySelector('button.dropdown-toggle').click();
    document.querySelectorAll('.dropdown-menu a')[5].click();
    '''
    driver.execute_script(dropdown_script)

    # 700개 항목이 로드되도록 충분한 시간 대기
    print("Waiting for the table to load with 700 items...")
    WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.XPATH, '//table/tbody/tr')))

# 다음 페이지로 이동
def go_to_next_page(driver):
    try:
        print("Clicking next page button...")
        next_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, 'a.datatable-pager-link-next'))
        )
        next_button.click()

        print("Waiting for next page to load...")
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//table/tbody/tr')))

    except Exception as e:
        print(f"Error navigating to next page: {e}")
        driver.quit()

# 할인 정보와 재고 정보 추출 함수
def process_inventory_data(cell):
    stock, expiration, discount_price = None, None, None  # 기본값 설정

    # 1. aria-label 속성에서 정보를 모두 추출
    stock_info = cell.get('aria-label')
    if stock_info:
        # 2. 재고 추출: 첫 번째 유통기한에서 재고를 추출
        expiration_stock_info = re.findall(r'(\d{2}/\d{2}/\d{4})\[(\d+)\]', stock_info)
        if expiration_stock_info:
            # 첫 번째 유통기한과 그에 대응하는 재고 추출
            expiration, stock = expiration_stock_info[0]
            stock = int(stock)  # 재고 값을 정수로 변환
        else:
            # 대괄호 안의 재고가 없을 때, 맨 앞에 숫자로 나오는 총 재고 추출
            leading_stock_info = re.match(r'^(\d+)', stock_info)
            if leading_stock_info:
                stock = int(leading_stock_info.group(1))  # 맨 앞의 총 재고 숫자를 추출

        # 3. 할인가 추출: $로 시작하는 숫자 추출
        discount_info = re.search(r'\$(\d+\.\d+)', stock_info)
        if discount_info:
            discount_price = float(discount_info.group(1))  # 할인가격을 숫자로 변환

            # 4. MD: 가 있는 경우 유통기한을 표시하지 않도록
            if 'MD:' in stock_info:
                expiration = None
        else:
            # 할인가가 없는 경우 유통기한을 표시하지 않고, 총 재고만 표시
            expiration = None
            leading_stock_info = re.match(r'^(\d+)', stock_info)  # 총 재고 추출
            if leading_stock_info:
                stock = int(leading_stock_info.group(1))  # 총 재고 값으로 설정

    # 5. 재고가 None이면 기본값 0으로 설정
    if stock is None:
        stock = 0

    return stock, discount_price, expiration

# BeautifulSoup을 사용한 데이터 스크래핑 및 Excel 파일로 저장
def scrape_data_and_save_to_excel(driver):
    # Excel 파일 생성
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Scraped Data"

    # 헤더 작성
    sheet.append(['Barcode', 'Description', 'CR_Stock', 'CR_Discount', 'CR_Expiration', 'NB_Stock', 'NB_Discount', 'NB_Expiration', 'IN_Stock', 'IN_Discount', 'IN_Expiration', 'BR_Stock', 'BR_Discount', 'BR_Expiration'])

    # 계속해서 페이지를 넘기며 데이터 스크래핑 (최대 4페이지)
    max_pages = 4  # 페이지 수 제한
    for page_number in range(1, max_pages + 1):
        print(f"Scraping page {page_number} data...")

        # 페이지 소스를 BeautifulSoup으로 파싱
        page_source = driver.page_source
        soup = BeautifulSoup(page_source, 'html.parser')

        # 테이블 행을 파싱하여 데이터를 추출
        rows = soup.select('table tbody tr')
        for row in rows:
            cells = row.find_all('td')

            # 필요한 데이터가 있는 경우 추출
            if len(cells) > 6:
                barcode = cells[0].get_text(strip=True)
                description = cells[1].get_text(strip=True)

                # 띄어쓰기 두 번을 한 번으로 교체
                description = re.sub(r'\s{2,}', ' ', description)

                # 바코드를 숫자 형식으로 변환
                if barcode.isdigit():
                    barcode = int(barcode)

                # 매장별 재고 및 할인 정보 추출
                cr_stock, cr_discount, cr_expiration = process_inventory_data(cells[2])
                nb_stock, nb_discount, nb_expiration = process_inventory_data(cells[3])
                in_stock, in_discount, in_expiration = process_inventory_data(cells[4])
                br_stock, br_discount, br_expiration = process_inventory_data(cells[5])

                # 엑셀 파일에 한 줄씩 추가
                sheet.append([barcode, description, cr_stock, cr_discount, cr_expiration, nb_stock, nb_discount, nb_expiration, in_stock, in_discount, in_expiration, br_stock, br_discount, br_expiration])

        if page_number < max_pages:
            go_to_next_page(driver)  # 다음 페이지로 이동

    # 현재 날짜 기준으로 파일 이름 작성 (yymmdd 형식)
    current_date = datetime.now().strftime('%y%m%d')
    file_name = f'Scraped_{current_date}.xlsx'

    # Excel 파일 저장
    workbook.save(file_name)
    print(f"Data successfully saved to '{file_name}'.")

# -----------------------
# 3. Shopify API 연동 코드
# -----------------------

# Shopify에서 모든 제품 가져오기
def get_shopify_products():
    products = []
    url = f"https://3abf38-d9.myshopify.com/admin/api/2024-07/products.json?limit=250"
    while url:
        response = requests.get(url, headers=headers)
        data = response.json()
        products.extend(data['products'])
        # 다음 페이지가 있는지 확인
        if 'Link' in response.headers:
            links = response.headers['Link']
            if 'rel="next"' in links:
                url = links.split(";")[0].strip("<>")
            else:
                url = None
        else:
            url = None
    return products

# 바코드와 상품 ID, 변형 ID 매핑 생성
def create_barcode_to_variant_map(products):
    barcode_to_variant = {}
    for product in products:
        for variant in product['variants']:
            barcode = variant.get('barcode')
            if barcode:
                barcode_to_variant[str(barcode)] = {
                    'product_id': product['id'],
                    'variant_id': variant['id'],
                    'variant': variant
                }
    return barcode_to_variant

# Shopify의 Location ID 가져오기
def get_shopify_locations():
    url = f"https://3abf38-d9.myshopify.com/admin/api/2024-07/locations.json"
    response = requests.get(url, headers=headers)
    data = response.json()
    locations = data['locations']
    return locations

# 재고 업데이트 함수
def update_inventory_level(inventory_item_id, location_id, available):
    url = f"https://3abf38-d9.myshopify.com/admin/api/2024-07/inventory_levels/set.json"
    payload = {
        "location_id": location_id,
        "inventory_item_id": inventory_item_id,
        "available": available
    }
    response = requests.post(url, headers=headers, json=payload)
    if response.status_code == 200:
        print(f"Inventory updated successfully for inventory_item_id: {inventory_item_id}")
    else:
        print(f"Failed to update inventory for inventory_item_id: {inventory_item_id}")
        print(response.text)

# 가격 및 할인 정보 업데이트 함수
def update_variant_price(variant_id, price, compare_at_price=None):
    url = f"https://3abf38-d9.myshopify.com/admin/api/2024-07/variants/{variant_id}.json"
    payload = {
        "variant": {
            "id": variant_id,
            "price": str(price)
        }
    }
    if compare_at_price:
        payload['variant']['compare_at_price'] = str(compare_at_price)
    else:
        payload['variant']['compare_at_price'] = None  # 할인가 없으면 제거

    response = requests.put(url, headers=headers, json=payload)
    if response.status_code == 200:
        print(f"Price updated successfully for variant_id: {variant_id}")
    else:
        print(f"Failed to update price for variant_id: {variant_id}")
        print(response.text)

# 스크래핑한 데이터 읽기
def read_scraped_data(file_name):
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active
    data = []
    # 헤더를 제외하고 데이터 읽기
    for row in sheet.iter_rows(min_row=2, values_only=True):
        barcode = str(row[0])
        description = row[1]
        # 각 매장의 재고 및 할인 정보
        stock_info = {
            'CR': {'stock': row[2], 'discount': row[3]},
            'NB': {'stock': row[5], 'discount': row[6]},
            'IN': {'stock': row[8], 'discount': row[9]},
            'BR': {'stock': row[11], 'discount': row[12]},
        }
        total_stock = sum([
            info['stock'] if info['stock'] is not None else 0
            for info in stock_info.values()
        ])
        # 최저 할인가 찾기
        discounts = [info['discount'] for info in stock_info.values() if info['discount'] is not None]
        min_discount = min(discounts) if discounts else None
        data.append({
            'barcode': barcode,
            'description': description,
            'total_stock': total_stock,
            'discount_price': min_discount
        })
    return data

# -----------------------
# 4. 메인 함수
# -----------------------

def main():
    # -----------------------
    # 4.1 스크래핑 실행
    # -----------------------
    print("Starting web scraping...")

    # ChromeDriver 설정
    chrome_service = Service(ChromeDriverManager().install())

    # 헤드리스 모드 활성화
    chrome_options = Options()
    chrome_options.add_argument('--headless')  # 헤드리스 모드 활성화
    chrome_options.add_argument('--disable-gpu')  # GPU 비활성화 (Windows에서 필요)
    chrome_options.add_argument('--window-size=1920x1080')  # 가상 브라우저 크기 설정

    # WebDriver 시작
    driver = webdriver.Chrome(service=chrome_service, options=chrome_options)

    try:
        login_and_navigate(driver)
        scrape_data_and_save_to_excel(driver)
    finally:
        driver.quit()

    # -----------------------
    # 4.2 Shopify 업데이트
    # -----------------------
    print("Updating Shopify with scraped data...")

    # Shopify에서 상품 정보 가져오기
    print("Fetching products from Shopify...")
    products = get_shopify_products()
    barcode_to_variant = create_barcode_to_variant_map(products)

    # Location ID 가져오기
    print("Fetching locations from Shopify...")
    locations = get_shopify_locations()
    if not locations:
        print("No locations found in Shopify.")
        return
    location_id = locations[0]['id']

    # 스크래핑한 데이터 읽기
    print("Reading scraped data...")
    current_date = datetime.now().strftime('%y%m%d')
    file_name = f'Scraped_{current_date}.xlsx'
    scraped_data = read_scraped_data(file_name)

    # 데이터 순회하며 업데이트
    for item in scraped_data:
        barcode = item['barcode']
        total_stock = item['total_stock']
        discount_price = item['discount_price']

        if barcode in barcode_to_variant:
            variant_info = barcode_to_variant[barcode]
            variant_id = variant_info['variant_id']
            inventory_item_id = variant_info['variant']['inventory_item_id']

            # 재고 업데이트
            update_inventory_level(inventory_item_id, location_id, total_stock)
            time.sleep(0.5)  # API 호출 제한을 고려하여 지연 시간 추가

            # 가격 업데이트 (할인가 적용)
            original_price = float(variant_info['variant']['price'])
            if discount_price is not None:
                update_variant_price(variant_id, price=discount_price, compare_at_price=original_price)
            else:
                # 할인가 없으면 원래 가격 유지하고 compare_at_price 제거
                update_variant_price(variant_id, price=original_price, compare_at_price=None)
            time.sleep(0.5)  # API 호출 제한을 고려하여 지연 시간 추가
        else:
            print(f"Barcode {barcode} not found in Shopify products.")

    print("Shopify update completed.")

if __name__ == "__main__":
    main()
