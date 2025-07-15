import openpyxl
from datetime import datetime

# 현재 날짜 가져오기 (yymmdd 형식)
current_date = datetime.now().strftime('%y%m%d')

# 파일 경로 설정
scraped_file = f'Scraped_{current_date}.xlsx'
itemlist_file = f'Itemlist_{current_date}.xlsx'

# Scraped_yymmdd.xlsx 로드 (시트 이름 지정)
print(f"Loading scraped file: {scraped_file}")
try:
    scraped_wb = openpyxl.load_workbook(scraped_file, read_only=False)
    scraped_sheet = scraped_wb.active  # 첫 번째 시트 사용
    print("Scraped file loaded successfully.")
except Exception as e:
    print(f"Error loading scraped file: {e}")
    scraped_sheet = None

# Itemlist_yymmdd.xlsx 로드 (시트 이름 지정)
print(f"Loading itemlist file: {itemlist_file}")
try:
    itemlist_wb = openpyxl.load_workbook(itemlist_file, read_only=False)
    itemlist_sheet = itemlist_wb.active  # 첫 번째 시트 사용
    print("Itemlist file loaded successfully.")
except Exception as e:
    print(f"Error loading itemlist file: {e}")
    itemlist_sheet = None

# Scraped 데이터 저장을 위한 사전
scraped_data = {}

# Scraped 파일에서 데이터 읽기 (Barcode, Title, BR_Stock, BR_Discount, BR_Expiration 추출)
if scraped_sheet:
    print("Reading scraped data...")
    try:
        for row in scraped_sheet.iter_rows(min_row=2, values_only=True):
            barcode = str(row[0])  # Scraped에서 BarCode
            title = row[1]  # Description
            br_stock = row[11] if row[11] is not None else 0  # BR_Stock (Scraped 데이터의 12번째 열)
            br_discount = row[12] if row[12] is not None else None  # BR_Discount (Scraped 데이터의 13번째 열)
            br_expiration = row[13] if row[13] is not None else None  # BR_Expiration (Scraped 데이터의 14번째 열)

            # 정확한 데이터 분리를 위한 추가 처리
            scraped_data[barcode] = {
                'title': title,
                'br_stock': br_stock,  # BR_Stock 사용
                'br_discount': br_discount,  # 할인 가격 데이터
                'br_expiration': br_expiration  # 유통기한 정보
            }
        print("Scraped data read successfully.")
    except Exception as e:
        print(f"Error reading scraped data: {e}")

# Itemlist 바코드 읽기
def get_itemlist_barcodes(itemlist_sheet):
    itemlist_barcodes = {}
    for row_idx in range(2, itemlist_sheet.max_row + 1):
        barcode = str(itemlist_sheet.cell(row=row_idx, column=14).value)  # N열의 Variant Barcode
        itemlist_barcodes[barcode] = row_idx
    return itemlist_barcodes

# 중복 바코드 확인 함수
def check_duplicate_barcodes(itemlist_sheet):
    barcodes_seen = set()
    duplicate_found = False
    for row_idx in range(2, itemlist_sheet.max_row + 1):
        barcode = str(itemlist_sheet.cell(row=row_idx, column=14).value)  # N열의 Variant Barcode
        if barcode in barcodes_seen:
            print(f"Duplicate barcode found: {barcode} at row {row_idx}")
            duplicate_found = True
            exit(1)  # 중복된 바코드가 있으면 작업 종료
        barcodes_seen.add(barcode)
    
    if not duplicate_found:
        print("중복된 바코드가 없습니다.")

# 1. 바코드를 비교하여 추가 및 삭제할 바코드 확인
def compare_barcodes(scraped_data, itemlist_barcodes):
    scraped_barcodes = set(scraped_data.keys())
    itemlist_barcodes_set = set(itemlist_barcodes.keys())
    
    # 추가해야 하는 바코드
    barcodes_to_add = scraped_barcodes - itemlist_barcodes_set
    
    # 삭제해야 하는 바코드
    barcodes_to_delete = itemlist_barcodes_set - scraped_barcodes
    
    return barcodes_to_add, barcodes_to_delete

# 2. 삭제할 바코드의 행을 삭제
def delete_barcodes(itemlist_barcodes, barcodes_to_delete, itemlist_sheet):
    for barcode in barcodes_to_delete:
        row_idx = itemlist_barcodes[barcode]
        print(f"Deleting row {row_idx} with barcode {barcode}")
        itemlist_sheet.delete_rows(row_idx)

# 3. 추가할 바코드를 Itemlist의 마지막에 추가
def add_barcodes(barcodes_to_add, scraped_data, itemlist_sheet):
    for barcode in barcodes_to_add:
        new_row = [
            scraped_data[barcode]['title'],  # Handle (A열: Title과 동일)
            scraped_data[barcode]['title'],  # Title (B열)
            None,  # Tags (C열)
            'TRUE',  # Published (D열)
            None,  # Variant SKU (E열)
            None,  # Variant Grams (F열)
            'shopify',  # Variant Inventory Tracker (G열)
            scraped_data[barcode]['br_stock'],  # Variant Inventory Qty (H열)
            'deny',  # Variant Inventory Policy (I열)
            'manual',  # Variant Fulfillment Service (J열)
            scraped_data[barcode]['br_discount'] if scraped_data[barcode]['br_discount'] is not None else None,  # Variant Price (K열)
            None,  # Variant Compare At Price (L열)
            'TRUE',  # Variant Taxable (M열)
            int(barcode),  # Variant Barcode (N열)
            None,  # Cost per item (O열)
            None,  # Status (P열)
            scraped_data[barcode]['br_expiration'],  # Expiration 추가 (Q열)
            None  # Info (R열)
        ]
        last_row = itemlist_sheet.max_row + 1
        print(f"Adding new barcode {barcode} at row {last_row}")
        itemlist_sheet.insert_rows(last_row)
        for col_idx, value in enumerate(new_row, start=1):
            itemlist_sheet.cell(row=last_row, column=col_idx, value=value)

# 4. 효율적으로 바코드를 재배열하는 함수
def reorder_barcodes(scraped_data, itemlist_sheet):
    """
    Scraped 파일의 바코드 순서에 맞게 Itemlist 파일의 바코드를 효율적으로 재배열.
    최소한의 이동만으로 최적화된 순서로 바코드를 재정렬.
    """
    # Scraped 데이터의 바코드 리스트
    scraped_barcodes = list(scraped_data.keys())
    
    # Itemlist의 바코드 리스트 (최신화된 상태)
    itemlist_barcodes = get_itemlist_barcodes(itemlist_sheet)
    
    # 필요할 때마다 이동할 위치 기록
    movements = []

    # Scraped 파일의 순서대로 진행하면서 Itemlist의 바코드를 이동
    for desired_position, barcode in enumerate(scraped_barcodes, start=2):  # 2행부터 시작
        # 바코드 위치 최신화
        itemlist_barcodes = get_itemlist_barcodes(itemlist_sheet)
        
        if barcode in itemlist_barcodes:
            current_position = itemlist_barcodes[barcode]
            
            # 바코드가 올바른 위치에 있지 않다면 이동
            if current_position != desired_position:
                movements.append((current_position, desired_position))
                print(f"Moving barcode {barcode} from row {current_position} to {desired_position}")

                # 이동 처리
                row_data = [cell.value for cell in itemlist_sheet[current_position]]
                itemlist_sheet.delete_rows(current_position)
                itemlist_sheet.insert_rows(desired_position)

                # 삭제된 행의 데이터를 새로운 위치에 다시 입력
                for col_idx, value in enumerate(row_data, start=1):
                    itemlist_sheet.cell(row=desired_position, column=col_idx, value=value)
                
                # 이동 후, 최신화
                itemlist_barcodes = get_itemlist_barcodes(itemlist_sheet)

    print("Reordering completed.")

# Variant SKU 업데이트 (E열)
def update_variant_sku(itemlist_sheet):
    """
    E열(Variant SKU)에 N열(Variant Barcode)와 같은 값을 수식으로 입력.
    """
    for row_idx in range(2, itemlist_sheet.max_row + 1):  # 2행부터 마지막 행까지
        formula = f"=N{row_idx}"  # 수식은 '='로 시작
        itemlist_sheet.cell(row=row_idx, column=5).value = formula  # E열에 수식 입력

# 6단계: Description, BR_Stock, BR_Discount 정보를 각각 Title, Variant Inventory Qty, Variant Price로 매칭
def update_existing_rows(scraped_data, itemlist_sheet):
    print("Matching and updating itemlist data...")
    try:
        for row_idx in range(2, itemlist_sheet.max_row + 1):
            variant_barcode = str(itemlist_sheet.cell(row=row_idx, column=14).value)  # N열의 Variant Barcode
            if variant_barcode in scraped_data:
                row = itemlist_sheet[row_idx]

                # Title 업데이트 (B열)
                if row[1].value != scraped_data[variant_barcode]['title']:
                    print(f"Title updated in row {row_idx} from {row[1].value} to {scraped_data[variant_barcode]['title']}")
                    row[1].value = scraped_data[variant_barcode]['title']

                # Variant Inventory Qty 업데이트 (H열, 인덱스는 8)
                row[7].value = scraped_data[variant_barcode]['br_stock']

                # Variant Price 업데이트 (K열, 인덱스는 11)
                if scraped_data[variant_barcode]['br_discount'] is not None:
                    row[10].value = scraped_data[variant_barcode]['br_discount']
                else:
                    row[10].value = row[11].value  # Compare At Price와 동일하게 설정
    except Exception as e:
        print(f"Error updating itemlist data: {e}")

# Status 열 (P열) 업데이트
def update_status_column(itemlist_sheet):
    print("Updating Status column (P열)...")
    try:
        for row_idx in range(2, itemlist_sheet.max_row + 1):
            status_formula = f'=IF(AND(Q{row_idx}="O", R{row_idx}="O"), "active", "archived")'
            itemlist_sheet.cell(row=row_idx, column=16).value = status_formula  # P열에 수식 입력 (인덱스는 16)
    except Exception as e:
        print(f"Error updating Status column: {e}")

# 코드 실행
# 1단계: Itemlist 파일 로드 후 중복 바코드 검사
check_duplicate_barcodes(itemlist_sheet)

# 2단계: 바코드를 비교하여 추가 및 삭제할 바코드 확인
itemlist_barcodes = get_itemlist_barcodes(itemlist_sheet)
barcodes_to_add, barcodes_to_delete = compare_barcodes(scraped_data, itemlist_barcodes)

# 3단계: 삭제할 바코드를 제거
delete_barcodes(itemlist_barcodes, barcodes_to_delete, itemlist_sheet)

# 4단계: 추가할 바코드를 가장 아래에 추가 + 바코드 중복 검사
add_barcodes(barcodes_to_add, scraped_data, itemlist_sheet)
check_duplicate_barcodes(itemlist_sheet)

# 5단계: 바코드 순서 정렬
reorder_barcodes(scraped_data, itemlist_sheet)

# 6단계: Description, BR_Stock, BR_Discount 정보를 각각 Title, Variant Inventory Qty, Variant Price로 매칭
update_existing_rows(scraped_data, itemlist_sheet)

# 7단계: Status 열 업데이트 (P열)
update_status_column(itemlist_sheet)

# 8단계: Variant SKU 업데이트
update_variant_sku(itemlist_sheet)

# 9단계: 업데이트된 Itemlist 파일 저장
if itemlist_sheet:
    updated_itemlist_file = f'Updated_Itemlist_{current_date}.xlsx'
    try:
        itemlist_wb.save(updated_itemlist_file)
        print(f"Updated Itemlist saved as '{updated_itemlist_file}'.")
    except Exception as e:
        print(f"Error saving updated itemlist file: {e}")
