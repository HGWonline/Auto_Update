import openpyxl
from datetime import datetime

# 현재 날짜 가져오기 (yymmdd 형식)
current_date = datetime.now().strftime('%y%m%d')

# 파일 경로 설정
scraped_file = f'Scraped_{current_date}.xlsx'
itemlist_file = f'Itemlist_{current_date}.xlsx'

# Excel 파일 로드
def load_excel(file_name):
    try:
        wb = openpyxl.load_workbook(file_name, read_only=False)
        sheet = wb.active
        print(f"Loaded file {file_name} successfully.")
        return wb, sheet
    except Exception as e:
        print(f"Error loading file {file_name}: {e}")
        return None, None

# 데이터 로드
scraped_wb, scraped_sheet = load_excel(scraped_file)
itemlist_wb, itemlist_sheet = load_excel(itemlist_file)

# Scraped 데이터 읽기
def read_scraped_data(scraped_sheet):
    scraped_data = {}
    for row in scraped_sheet.iter_rows(min_row=2, values_only=True):
        barcode = str(row[0])
        scraped_data[barcode] = {
            'title': row[1],
            'ca_stock': row[2],
            'ca_discount': row[3],
            'ca_expiration': row[4],
            'nb_stock': row[5],
            'nb_discount': row[6],
            'nb_expiration': row[7],
            'in_stock': row[8],
            'in_discount': row[9],
            'in_expiration': row[10],
            'br_stock': row[11],
            'br_discount': row[12],
            'br_expiration': row[13]
        }
    return scraped_data

# Itemlist 바코드 읽기
def get_itemlist_barcodes(itemlist_sheet):
    itemlist_barcodes = {}
    for row_idx in range(2, itemlist_sheet.max_row + 1, 4):
        barcode = str(itemlist_sheet.cell(row=row_idx, column=6).value)
        if barcode:
            itemlist_barcodes[barcode] = row_idx
    return itemlist_barcodes

# 중복 바코드 확인
def check_duplicate_barcodes(itemlist_sheet):
    barcodes_seen = set()
    for row_idx in range(2, itemlist_sheet.max_row + 1):
        barcode = str(itemlist_sheet.cell(row=row_idx, column=7).value)
        if barcode and barcode in barcodes_seen:
            print(f"Duplicate barcode found: {barcode} at row {row_idx}")
            return True
        barcodes_seen.add(barcode)
    return False

# 바코드 비교
def compare_barcodes(scraped_data, itemlist_barcodes):
    scraped_barcodes = set(scraped_data.keys())
    itemlist_barcodes_set = set(itemlist_barcodes.keys())
    return scraped_barcodes - itemlist_barcodes_set, itemlist_barcodes_set - scraped_barcodes

# 바코드 삭제
def delete_barcodes(barcodes_to_delete, itemlist_barcodes, itemlist_sheet):
    for barcode in barcodes_to_delete:
        row_idx = itemlist_barcodes[barcode]
        print(f"Deleting rows {row_idx} to {row_idx+3} with barcode {barcode}")
        itemlist_sheet.delete_rows(row_idx, amount=4)

# 바코드 추가
def add_barcodes(barcodes_to_add, scraped_data, itemlist_sheet):
    store_names = ['Booragoon', 'Carousel', 'Northbridge', 'Innaloo']

    for barcode in barcodes_to_add:
        new_rows = [[scraped_data[barcode]['title'], scraped_data[barcode]['title'], None, None, 'Location', store_names[i],
                     barcode if i == 0 else '',  # Booragoon에만 바코드 기록
                     scraped_data[barcode]['br_stock'] if i == 0 else '',  # Booragoon에만 재고 기록
                     scraped_data[barcode]['br_discount'] if i == 0 else '', None, None, None, None, None, None]
                    for i in range(4)]

        last_row = itemlist_sheet.max_row + 1
        print(f"Adding new barcode {barcode} at rows {last_row} to {last_row + 3}")

        for i, new_row in enumerate(new_rows):
            itemlist_sheet.insert_rows(last_row + i)
            for col_idx, value in enumerate(new_row, start=1):
                itemlist_sheet.cell(row=last_row + i, column=col_idx, value=value)

# 바코드 재정렬 (Handle 값이 같은 4개의 행을 하나의 세트로 함께 이동)
def reorder_barcodes(scraped_data, itemlist_sheet):
    scraped_barcodes = list(scraped_data.keys())
    itemlist_barcodes = get_itemlist_barcodes(itemlist_sheet)

    sorted_data = []

    for barcode in scraped_barcodes:
        if barcode in itemlist_barcodes:
            current_position = itemlist_barcodes[barcode]

            rows_data = []
            for i in range(4):
                row_data = [cell.value for cell in itemlist_sheet[current_position + i]]
                rows_data.append(row_data)

            sorted_data.extend(rows_data)

    # 기존 데이터를 삭제하고 재정렬된 데이터를 삽입
    itemlist_sheet.delete_rows(2, itemlist_sheet.max_row)

    for row_idx, row_data in enumerate(sorted_data, start=2):
        for col_idx, cell_value in enumerate(row_data, start=1):
            itemlist_sheet.cell(row=row_idx, column=col_idx, value=cell_value)

    print("Reordering completed.")

# 매장 이름 확인 및 수정
def ensure_store_names(itemlist_sheet):
    store_names = ['Booragoon', 'Carousel', 'Northbridge', 'Innaloo']
    for row_idx in range(2, itemlist_sheet.max_row + 1, 4):
        for i, store_name in enumerate(store_names):
            cell_value = itemlist_sheet.cell(row=row_idx + i, column=5).value
            if cell_value != store_name:
                print(f"Missing store name at row {row_idx + i}, writing '{store_name}'")
                itemlist_sheet.cell(row=row_idx + i, column=5).value = store_name

# Itemlist 업데이트
def update_itemlist(scraped_data, itemlist_sheet):
    stock_keys = ['br_stock', 'ca_stock', 'nb_stock', 'in_stock']
    discount_keys = ['br_discount', 'ca_discount', 'nb_discount', 'in_discount']
    expiration_keys = ['br_expiration', 'ca_expiration', 'nb_expiration', 'in_expiration']
    store_names = ['Booragoon', 'Carousel', 'Northbridge', 'Innaloo']

    for row_idx in range(2, itemlist_sheet.max_row + 1, 4):
        handle = itemlist_sheet.cell(row=row_idx, column=6).value  # Booragoon의 바코드로 매칭

        if handle in scraped_data:
            product_data = scraped_data[handle]

            for i in range(4):
                itemlist_row_idx = row_idx + i

                # Variant Barcode, Title 업데이트 (Booragoon에만 처리)
                if i == 0:
                    itemlist_sheet.cell(row=itemlist_row_idx, column=6).value = handle  # Booragoon의 바코드
                    itemlist_sheet.cell(row=itemlist_row_idx, column=2).value = product_data['title']

                itemlist_sheet.cell(row=itemlist_row_idx, column=1).value = product_data['title']
                itemlist_sheet.cell(row=itemlist_row_idx, column=5).value = store_names[i]

                # 재고 정보 업데이트 (H열)
                itemlist_sheet.cell(row=itemlist_row_idx, column=8).value = product_data.get(stock_keys[i], '')

                # 할인 가격 정보 업데이트 (I열)
                itemlist_sheet.cell(row=itemlist_row_idx, column=9).value = product_data.get(discount_keys[i], '')

                # 유통기한 정보 업데이트 (V, W, X, Y열)
                if i == 0:
                    itemlist_sheet.cell(row=itemlist_row_idx, column=22).value = product_data.get('br_expiration', '')
                    itemlist_sheet.cell(row=itemlist_row_idx, column=23).value = product_data.get('ca_expiration', '')
                    itemlist_sheet.cell(row=itemlist_row_idx, column=24).value = product_data.get('nb_expiration', '')
                    itemlist_sheet.cell(row=itemlist_row_idx, column=25).value = product_data.get('in_expiration', '')

# Status 열 업데이트
def update_status_column(itemlist_sheet):
    for row_idx in range(2, itemlist_sheet.max_row + 1):
        status_formula = f'=IF(AND(S{row_idx}="O", T{row_idx}="O"), "active", "archived")'
        itemlist_sheet.cell(row=row_idx, column=18).value = status_formula
        for i in range(1, 4):
            itemlist_sheet.cell(row=row_idx + i, column=18).value = ''

# 실행 부분
if not check_duplicate_barcodes(itemlist_sheet):
    scraped_data = read_scraped_data(scraped_sheet)
    itemlist_barcodes = get_itemlist_barcodes(itemlist_sheet)
    
    barcodes_to_add, barcodes_to_delete = compare_barcodes(scraped_data, itemlist_barcodes)

    delete_barcodes(barcodes_to_delete, itemlist_barcodes, itemlist_sheet)
    add_barcodes(barcodes_to_add, scraped_data, itemlist_sheet)
    reorder_barcodes(scraped_data, itemlist_sheet)
    ensure_store_names(itemlist_sheet)
    update_itemlist(scraped_data, itemlist_sheet)
    update_status_column(itemlist_sheet)

    updated_itemlist_file = f'Updated_Itemlist_{current_date}.xlsx'
    itemlist_wb.save(updated_itemlist_file)
    print(f"Updated Itemlist saved as '{updated_itemlist_file}'.")
