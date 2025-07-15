import openpyxl
import csv
from datetime import datetime
import os

# 현재 날짜 (yymmdd)
current_date = datetime.now().strftime('%y%m%d')

# 파일명 설정
itemlist_file = f'Itemlist_{current_date}.xlsx'
old_promotion_file = f'Promotion_{current_date}.csv'
updated_promotion_file = f'Updated_Promotion_{current_date}.csv'
log_file = f'Log_Promotion_{current_date}.csv'

# Itemlist 엑셀 로드
try:
    itemlist_wb = openpyxl.load_workbook(itemlist_file, read_only=True, data_only=True)
    itemlist_sheet = itemlist_wb.active
except Exception as e:
    print(f"Error loading Itemlist file: {e}")
    itemlist_sheet = None

# Promotion 파일 헤더
headers = ['Name', 'Rate', 'Price', 'End date', 'Location', 'Online']
promotion_data = []

if itemlist_sheet:
    # 최대 3000행까지만 처리 (필요없으면 제거 가능)
    max_row_to_process = min(itemlist_sheet.max_row, 3000)
    rows = itemlist_sheet.iter_rows(min_row=2, max_row=max_row_to_process, values_only=True)

    for idx, row in enumerate(rows, start=2):
        # 인덱스: 
        # A(0): Handle, B(1): Title, ...
        # K(10): Variant Price, L(11): Variant Compare At Price
        # P(15): Status, T(19): Expiration_date
        # U(20), V(21), W(22), X(23): 매장 재고 상태
        title = row[1]
        variant_price = row[10]
        variant_compare_price = row[11]
        status = row[15]
        expiration_cell = row[19]
        stock_cr = row[20]
        stock_nb = row[21]
        stock_in = row[22]
        stock_br = row[23]

        if variant_price is None or variant_compare_price is None:
            continue

        if variant_price != variant_compare_price:
            # Name
            name = title if title else ""

            # Price
            price = variant_price if variant_price else 0

            # Rate 계산: (L-K)/K * 100 이 아닌 기존 코드에서 (L-K)/L * 100로 되어 있었음
            # 문제에서 새로 수정된 부분 없으니 기존 로직 유지.
            # 문제에서 제시한 (L-K)/K*100 으로 다시 수정할 필요가 있음. (유저 요구사항 재확인)
            # 요구사항 5번: Rate = (L-K)/K*100
            try:
                k = float(variant_price)
                l = float(variant_compare_price)
                if k != 0:
                    rate_val = (l - k) / k * 100
                    # 정수 판별
                    if rate_val.is_integer():
                        rate_str = f"{int(rate_val)}%"
                    else:
                        rate_str = "0"
                else:
                    rate_str = "0"
            except:
                rate_str = "0"

            # End date 추출
            end_date = ""
            if expiration_cell:
                str_val = str(expiration_cell)
                if "Expiration date:" in str_val:
                    end_date = str_val.replace("Expiration date:", "").strip()
                else:
                    end_date = str_val.strip()

            # Location
            store_mapping = [('CA', stock_cr), ('NB', stock_nb), ('IN', stock_in), ('BR', stock_br)]
            stores = [code for code, val in store_mapping if val is not None and "Out of stock" not in str(val)]
            if len(stores) == 4:
                location = "ALL"
            else:
                location = ", ".join(stores) if stores else ""

            # Online
            online = "O" if status == "active" else "X"

            promotion_data.append([name, rate_str, price, end_date, location, online])

        # 진행 상황 출력(옵션)
        if (idx % 500) == 0:
            print(f"Processed {idx} rows...")

# 기존 Promotion 파일 로드 (비교 위해)
old_promotion_data = []
if os.path.exists(old_promotion_file):
    with open(old_promotion_file, 'r', encoding='utf-8-sig', newline='') as f:
        reader = csv.reader(f)
        old_headers = next(reader, None)
        for row in reader:
            if len(row) == len(headers):
                old_promotion_data.append(row)
else:
    # 기존 파일이 없으면 빈 리스트로 처리
    old_promotion_data = []

# 새로운 Promotion 파일 (Updated_Promotion_yymmdd.csv) 저장
with open(updated_promotion_file, 'w', newline='', encoding='utf-8-sig') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(headers)
    writer.writerows(promotion_data)

print(f"Updated Promotion data saved to {updated_promotion_file}")

# 변경점 비교
# Name 필드를 기준으로 old/new를 매핑
def list_to_dict(data_list, key_index=0):
    d = {}
    for r in data_list:
        key = r[key_index]
        d[key] = r
    return d

old_dict = list_to_dict(old_promotion_data, 0)  # old Name -> old row
new_dict = list_to_dict(promotion_data, 0)      # new Name -> new row

# 변경사항 기록
# 로그 파일 헤더: 
# ChangeType, Name, OldRate, NewRate, OldPrice, NewPrice, OldEndDate, NewEndDate, OldLocation, NewLocation, OldOnline, NewOnline
log_headers = ["ChangeType", "Name", "OldRate", "NewRate", "OldPrice", "NewPrice",
               "OldEndDate", "NewEndDate", "OldLocation", "NewLocation", "OldOnline", "NewOnline"]
log_rows = []

# Deleted: old에는 있고 new에는 없는 경우
for old_name, old_row in old_dict.items():
    if old_name not in new_dict:
        # 삭제된 행
        # old_row: [Name, Rate, Price, End date, Location, Online]
        log_rows.append(["Deleted", old_row[0],
                         old_row[1], "", old_row[2], "", old_row[3], "", old_row[4], "", old_row[5], ""])

# Added: new에는 있고 old에는 없는 경우
for new_name, new_row in new_dict.items():
    if new_name not in old_dict:
        # 추가된 행
        # new_row: [Name, Rate, Price, End date, Location, Online]
        log_rows.append(["Added", new_row[0],
                         "", new_row[1], "", new_row[2], "", new_row[3], "", new_row[4], "", new_row[5]])

# Modified: 둘 다 있는데 내용이 다른 경우
for name, new_row in new_dict.items():
    if name in old_dict:
        old_row = old_dict[name]
        # 비교 (Name 제외, 인덱스 1~5)
        # [Name, Rate, Price, End date, Location, Online]
        changed = False
        for i in range(1, len(headers)):
            if str(old_row[i]) != str(new_row[i]):
                changed = True
                break
        if changed:
            log_rows.append(["Modified", name,
                             old_row[1], new_row[1],
                             old_row[2], new_row[2],
                             old_row[3], new_row[3],
                             old_row[4], new_row[4],
                             old_row[5], new_row[5]])

# 로그 파일 작성
with open(log_file, 'w', newline='', encoding='utf-8-sig') as f:
    writer = csv.writer(f)
    writer.writerow(log_headers)
    writer.writerows(log_rows)

print(f"Log of changes saved to {log_file}")
