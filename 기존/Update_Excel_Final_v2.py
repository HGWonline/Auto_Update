import openpyxl
from datetime import datetime

# 현재 날짜 가져오기 (yymmdd 형식)
current_date = datetime.now().strftime('%y%m%d')

# 파일 경로 설정
scraped_file = f'Scraped_{current_date}.xlsx'
itemlist_file = f'Itemlist_{current_date}.xlsx'
scraped_m_file = f'ScrapedM_{current_date}.xlsx'  # 추가된 ScrapedM 파일

# Scraped_yymmdd.xlsx 로드 (시트 이름 지정)
print(f"Loading scraped file: {scraped_file}")
try:
    scraped_wb = openpyxl.load_workbook(scraped_file, read_only=False)
    scraped_sheet = scraped_wb.active  # 첫 번째 시트 사용
    print("Scraped file loaded successfully.")
except Exception as e:
    print(f"Error loading scraped file: {e}")
    scraped_sheet = None

# ScrapedM_yymmdd.xlsx 로드 (시트 이름 지정)
print(f"Loading scrapedM file: {scraped_m_file}")
try:
    scraped_m_wb = openpyxl.load_workbook(scraped_m_file, read_only=False)
    scraped_m_sheet = scraped_m_wb.active  # 첫 번째 시트 사용
    print("ScrapedM file loaded successfully.")
except Exception as e:
    print(f"Error loading scrapedM file: {e}")
    scraped_m_sheet = None

# Itemlist_yymmdd.xlsx 로드 (시트 이름 지정)
print(f"Loading itemlist file: {itemlist_file}")
try:
    itemlist_wb = openpyxl.load_workbook(itemlist_file, read_only=False)
    itemlist_sheet = itemlist_wb.active  # 첫 번째 시트 사용
    print("Itemlist file loaded successfully.")
except Exception as e:
    print(f"Error loading itemlist file: {e}")
    itemlist_sheet = None

# Scraped 데이터 저장을 위한 사전
scraped_data = {}

# Scraped 파일에서 데이터 읽기
if scraped_sheet:
    print("Reading scraped data...")
    try:
        for row in scraped_sheet.iter_rows(min_row=2, values_only=True):
            barcode = str(row[0])  # BarCode
            title = row[1]  # Description

            # 4개 매장의 재고 정보 가져오기
            cr_stock = row[2] if row[2] is not None else 0  # CR_Stock
            cr_discount = row[3] if row[3] is not None else None  # CR_Discount
            cr_expiration = row[4] if row[4] else None  # CR_Expiration

            nb_stock = row[5] if row[5] is not None else 0  # NB_Stock
            nb_discount = row[6] if row[6] is not None else None  # NB_Discount
            nb_expiration = row[7] if row[7] else None  # NB_Expiration

            in_stock = row[8] if row[8] is not None else 0  # IN_Stock
            in_discount = row[9] if row[9] is not None else None  # IN_Discount
            in_expiration = row[10] if row[10] else None  # IN_Expiration

            br_stock = row[11] if row[11] is not None else 0  # BR_Stock
            br_discount = row[12] if row[12] is not None else None  # BR_Discount
            br_expiration = row[13] if row[13] else None  # BR_Expiration

            # 총 재고 합산
            total_stock = cr_stock + nb_stock + in_stock + br_stock

            # 유통기한이 있는 매장의 재고 및 할인 정보 가져오기
            expiration_infos = []
            if cr_expiration:
                expiration_infos.append((cr_expiration, cr_stock, cr_discount))
            if nb_expiration:
                expiration_infos.append((nb_expiration, nb_stock, nb_discount))
            if in_expiration:
                expiration_infos.append((in_expiration, in_stock, in_discount))
            if br_expiration:
                expiration_infos.append((br_expiration, br_stock, br_discount))

            if expiration_infos:
                # 유통기한 순서대로 정렬
                expiration_infos.sort(key=lambda x: x[0])

                # 유효한 재고가 있는 가장 빠른 유통기한 찾기
                processed_dates = set()
                found_valid_stock = False

                for exp_date, stock, discount in expiration_infos:
                    if exp_date in processed_dates:
                        continue
                    processed_dates.add(exp_date)

                    # 해당 유통기한의 재고 합산
                    exp_total_stock = sum(s for ed, s, d in expiration_infos if ed == exp_date)

                    if exp_total_stock > 0:
                        # 재고가 있으면 해당 유통기한 사용
                        total_stock = exp_total_stock
                        expiration_date = exp_date
                        # 할인 가격 설정
                        discounts = [d for ed, s, d in expiration_infos if ed == exp_date and d is not None]
                        discount_price = discounts[0] if discounts else None
                        found_valid_stock = True
                        break

                if not found_valid_stock:
                    # 모든 유통기한에 대해 재고가 0인 경우
                    total_stock = 0
                    expiration_date = None
                    discount_price = None
            else:
                # 유통기한이 없으면 None
                expiration_date = None
                discount_price = None

            scraped_data[barcode] = {
                'title': title,
                'total_stock': total_stock,
                'discount_price': discount_price,
                'expiration_date': expiration_date
            }
        print("Scraped data read successfully.")
    except Exception as e:
        print(f"Error reading scraped data: {e}")

# ScrapedM 데이터 읽기
def read_scraped_m_data(scraped_m_sheet):
    scraped_m_data = {}
    for row in scraped_m_sheet.iter_rows(min_row=2, values_only=True):
        barcode = str(row[1])  # Barcode
        wholesaler_price = row[2]  # wholesalerPrice
        retail_price = row[3]  # retailPrice

        scraped_m_data[barcode] = {
            'wholesaler_price': wholesaler_price,
            'retail_price': retail_price
        }
    return scraped_m_data

# Itemlist 바코드 읽기
def get_itemlist_barcodes(itemlist_sheet):
    itemlist_barcodes = {}
    for row_idx in range(2, itemlist_sheet.max_row + 1):
        barcode = str(itemlist_sheet.cell(row=row_idx, column=14).value)  # N열의 Variant Barcode
        itemlist_barcodes[barcode] = row_idx
    return itemlist_barcodes

# 중복 바코드 확인 함수
def check_duplicate_barcodes(itemlist_sheet):
    barcodes_seen = set()
    duplicate_found = False
    for row_idx in range(2, itemlist_sheet.max_row + 1):
        barcode = str(itemlist_sheet.cell(row=row_idx, column=14).value)  # N열의 Variant Barcode
        if barcode in barcodes_seen:
            print(f"Duplicate barcode found: {barcode} at row {row_idx}")
            duplicate_found = True
            exit(1)  # 중복된 바코드가 있으면 작업 종료
        barcodes_seen.add(barcode)

    if not duplicate_found:
        print("중복된 바코드가 없습니다.")

# 1. 바코드를 비교하여 추가 및 삭제할 바코드 확인
def compare_barcodes(scraped_data, itemlist_barcodes):
    scraped_barcodes = set(scraped_data.keys())
    itemlist_barcodes_set = set(itemlist_barcodes.keys())

    # 추가해야 하는 바코드
    barcodes_to_add = scraped_barcodes - itemlist_barcodes_set

    # 삭제해야 하는 바코드
    barcodes_to_delete = itemlist_barcodes_set - scraped_barcodes

    return barcodes_to_add, barcodes_to_delete

# 2. 삭제할 바코드의 행을 삭제 (뒤에서부터 삭제)
def delete_barcodes(itemlist_barcodes, barcodes_to_delete, itemlist_sheet):
    rows_to_delete = [itemlist_barcodes[barcode] for barcode in barcodes_to_delete]
    rows_to_delete.sort(reverse=True)  # 역순으로 정렬하여 뒤에서부터 삭제
    for row_idx in rows_to_delete:
        barcode = str(itemlist_sheet.cell(row=row_idx, column=14).value)
        print(f"Deleting row {row_idx} with barcode {barcode}")
        itemlist_sheet.delete_rows(row_idx)

# 3. 추가할 바코드를 Itemlist의 마지막에 추가
def add_barcodes(barcodes_to_add, scraped_data, itemlist_sheet):
    for barcode in barcodes_to_add:
        new_row = [
            scraped_data[barcode]['title'],  # Handle (A열)
            scraped_data[barcode]['title'],  # Title (B열)
            None,  # Tags (C열)
            'TRUE',  # Published (D열)
            None,  # Variant SKU (E열)
            None,  # Variant Grams (F열)
            'shopify',  # Variant Inventory Tracker (G열)
            scraped_data[barcode]['total_stock'],  # Variant Inventory Qty (H열)
            'deny',  # Variant Inventory Policy (I열)
            'manual',  # Variant Fulfillment Service (J열)
            None,  # Variant Price (K열)
            None,  # Variant Compare At Price (L열)
            'TRUE',  # Variant Taxable (M열)
            int(barcode),  # Variant Barcode (N열)
            None,  # Cost per item (O열)
            None,  # Status (P열)
            None,  # Gift Card (Q열)
            None,  # SEO Title (R열)
            None,  # SEO Description (S열)
            None  # metafield.custom.expiration_date (T열)
        ]
        last_row = itemlist_sheet.max_row + 1
        print(f"Adding new barcode {barcode} at row {last_row}")
        for col_idx, value in enumerate(new_row, start=1):
            itemlist_sheet.cell(row=last_row, column=col_idx, value=value)

# swap_rows 함수 정의
def swap_rows(sheet, row_idx1, row_idx2):
    """
    sheet에서 row_idx1 행과 row_idx2 행의 셀 값을 교환합니다.
    셀의 값뿐만 아니라 스타일과 속성도 함께 교환합니다.
    """
    max_col = sheet.max_column
    for col_idx in range(1, max_col + 1):
        cell1 = sheet.cell(row=row_idx1, column=col_idx)
        cell2 = sheet.cell(row=row_idx2, column=col_idx)

        # 셀 값 교환
        cell1_value = cell1.value
        cell1.value = cell2.value
        cell2.value = cell1_value

        # 데이터 유형 교환
        cell1_data_type = cell1.data_type
        cell1.data_type = cell2.data_type
        cell2.data_type = cell1_data_type

        # 스타일 교환
        cell1_font = cell1.font
        cell1.font = cell2.font
        cell2.font = cell1_font

        cell1_fill = cell1.fill
        cell1.fill = cell2.fill
        cell2.fill = cell1_fill

        cell1_border = cell1.border
        cell1.border = cell2.border
        cell2.border = cell1_border

        cell1_alignment = cell1.alignment
        cell1.alignment = cell2.alignment
        cell2.alignment = cell1_alignment

        cell1_number_format = cell1.number_format
        cell1.number_format = cell2.number_format
        cell2.number_format = cell1_number_format

        cell1_protection = cell1.protection
        cell1.protection = cell2.protection
        cell2.protection = cell1_protection

        # 주석 교환
        cell1_comment = cell1.comment
        cell1.comment = cell2.comment
        cell2.comment = cell1_comment

        # 하이퍼링크 교환
        cell1_hyperlink = cell1.hyperlink
        cell1.hyperlink = cell2.hyperlink
        cell2.hyperlink = cell1_hyperlink

# 4. 바코드를 재배열하는 함수
def reorder_barcodes(scraped_data, itemlist_sheet):
    """
    Scraped 파일의 바코드 순서에 맞게 Itemlist 파일의 바코드를 재배열.
    기존의 데이터를 보존하면서 행의 순서를 변경합니다.
    """
    print("Reordering rows based on barcode sequence...")
    # Scraped 데이터의 바코드 리스트
    scraped_barcodes = list(scraped_data.keys())

    # Itemlist의 바코드와 행 번호 매핑
    itemlist_barcodes = get_itemlist_barcodes(itemlist_sheet)

    # 현재 행 위치와 목표 행 위치 매핑
    barcode_to_current_row = {barcode: row_idx for barcode, row_idx in itemlist_barcodes.items()}
    barcode_to_target_row = {}
    target_row = 2  # 헤더 이후 첫 번째 행부터 시작
    for barcode in scraped_barcodes:
        if barcode in barcode_to_current_row:
            barcode_to_target_row[barcode] = target_row
            target_row += 1
        else:
            print(f"Barcode {barcode} not found in itemlist during reordering.")

    # 이미 처리한 행을 추적하기 위한 집합
    processed_rows = set()

    # 행 이동 작업 수행
    for barcode, target_row_idx in barcode_to_target_row.items():
        current_row_idx = barcode_to_current_row[barcode]
        if current_row_idx != target_row_idx and current_row_idx not in processed_rows:
            # 현재 위치와 목표 위치가 다른 경우 행 교환
            swap_rows(itemlist_sheet, current_row_idx, target_row_idx)
            processed_rows.update({current_row_idx, target_row_idx})
            # 교환 후, 위치 정보 업데이트
            barcode_to_current_row[barcode] = target_row_idx
            # 교환된 행의 바코드도 업데이트
            swapped_barcode = itemlist_sheet.cell(row=current_row_idx, column=14).value
            if swapped_barcode:
                barcode_to_current_row[str(swapped_barcode)] = current_row_idx
    print("Reordering completed.")

# 5. ScrapedM 파일의 데이터를 업데이트
def update_prices(scraped_m_data, itemlist_sheet):
    itemlist_barcodes = get_itemlist_barcodes(itemlist_sheet)  # 최신화된 바코드 맵
    for barcode, row_idx in itemlist_barcodes.items():
        if barcode in scraped_m_data:
            data = scraped_m_data[barcode]
            # retailPrice를 K열과 L열에 업데이트
            itemlist_sheet.cell(row=row_idx, column=11).value = data['retail_price']  # K열
            itemlist_sheet.cell(row=row_idx, column=12).value = data['retail_price']  # L열
            # wholesalerPrice를 O열에 업데이트
            itemlist_sheet.cell(row=row_idx, column=15).value = data['wholesaler_price']  # O열
            print(f"Updated prices for barcode {barcode}: retailPrice={data['retail_price']}, wholesalerPrice={data['wholesaler_price']}")
        else:
            # ScrapedM에 바코드가 없을 경우 기본값 설정
            itemlist_sheet.cell(row=row_idx, column=11).value = 0  # K열
            itemlist_sheet.cell(row=row_idx, column=12).value = 0  # L열
            itemlist_sheet.cell(row=row_idx, column=15).value = 0  # O열
            print(f"Barcode {barcode} not found in ScrapedM. Prices set to 0.")

# 6. Scraped 데이터를 기반으로 업데이트
def update_existing_rows(scraped_data, itemlist_sheet):
    print("Matching and updating itemlist data...")
    try:
        for row_idx in range(2, itemlist_sheet.max_row + 1):
            variant_barcode = str(itemlist_sheet.cell(row=row_idx, column=14).value)
            if variant_barcode in scraped_data:
                # 필요한 열만 업데이트
                scraped_info = scraped_data[variant_barcode]

                # Title 업데이트 (B열)
                itemlist_sheet.cell(row=row_idx, column=2).value = scraped_info['title']

                # Variant Inventory Qty 업데이트 (H열)
                itemlist_sheet.cell(row=row_idx, column=8).value = scraped_info['total_stock']
    except Exception as e:
        print(f"Error updating itemlist data: {e}")

# 7. discount_price를 K열에 업데이트
def update_discount(scraped_data, itemlist_barcodes, itemlist_sheet):
    for barcode, data in scraped_data.items():
        if barcode in itemlist_barcodes:
            row_idx = itemlist_barcodes[barcode]
            # discount_price가 있는 경우에만 K열에 덮어쓰기
            if 'discount_price' in data and data['discount_price'] is not None:
                itemlist_sheet.cell(row=row_idx, column=11).value = data['discount_price']  # K열
                print(f"Updated discount price for barcode {barcode}: {data['discount_price']}")

# 8. expiration_date를 T열에 업데이트
def update_expiration(scraped_data, itemlist_barcodes, itemlist_sheet):
    for barcode, data in scraped_data.items():
        if barcode in itemlist_barcodes:
            row_idx = itemlist_barcodes[barcode]
            expiration_date = data['expiration_date']

            # Expiration Date가 있는 경우
            if expiration_date:
                expiration_text = f"Expiration date: {expiration_date}"
                itemlist_sheet.cell(row=row_idx, column=20).value = expiration_text  # T열은 20번째 열
            else:
                itemlist_sheet.cell(row=row_idx, column=20).value = None  # 정보가 없으면 빈칸 유지

# 9. Status 열 업데이트 (P열)
def update_status_column(itemlist_sheet):
    print("Updating Status column (P열)...")
    try:
        for row_idx in range(2, itemlist_sheet.max_row + 1):
            status_formula = f'=IF(AND(Q{row_idx}="O", R{row_idx}="O"), "active", "archived")'
            itemlist_sheet.cell(row=row_idx, column=16).value = status_formula  # P열에 수식 입력
    except Exception as e:
        print(f"Error updating Status column: {e}")

# 10. Variant SKU 업데이트 (E열)
def update_variant_sku(itemlist_sheet):
    for row_idx in range(2, itemlist_sheet.max_row + 1):
        barcode = itemlist_sheet.cell(row=row_idx, column=14).value  # N열의 바코드
        itemlist_sheet.cell(row=row_idx, column=5).value = barcode  # E열에 바코드 입력

# 11. 재고가 0인 경우 가격 조정
def adjust_prices_for_zero_stock(itemlist_sheet):
    print("Adjusting prices for items with zero stock...")
    try:
        for row_idx in range(2, itemlist_sheet.max_row + 1):
            inventory_qty = itemlist_sheet.cell(row=row_idx, column=8).value  # H열
            if inventory_qty is None or inventory_qty == 0 or inventory_qty == '0':
                variant_compare_at_price = itemlist_sheet.cell(row=row_idx, column=12).value  # L열
                itemlist_sheet.cell(row=row_idx, column=11).value = variant_compare_at_price  # K열
    except Exception as e:
        print(f"Error adjusting prices: {e}")

# 코드 실행
# 1단계: 중복 바코드 검사
check_duplicate_barcodes(itemlist_sheet)

# 2단계: 바코드 비교하여 추가 및 삭제할 바코드 확인
itemlist_barcodes = get_itemlist_barcodes(itemlist_sheet)
barcodes_to_add, barcodes_to_delete = compare_barcodes(scraped_data, itemlist_barcodes)

# 3단계: 삭제할 바코드 제거
delete_barcodes(itemlist_barcodes, barcodes_to_delete, itemlist_sheet)

# 4단계: 추가할 바코드를 추가하고 중복 검사
add_barcodes(barcodes_to_add, scraped_data, itemlist_sheet)
check_duplicate_barcodes(itemlist_sheet)

# 5단계: 바코드 순서 정렬
itemlist_barcodes = get_itemlist_barcodes(itemlist_sheet)
reorder_barcodes(scraped_data, itemlist_sheet)

# 6단계: 데이터 업데이트
update_existing_rows(scraped_data, itemlist_sheet)

# 7단계: 가격 업데이트
scraped_m_data = read_scraped_m_data(scraped_m_sheet)
update_prices(scraped_m_data, itemlist_sheet)

# 8단계: 할인 가격 업데이트
itemlist_barcodes = get_itemlist_barcodes(itemlist_sheet)
update_discount(scraped_data, itemlist_barcodes, itemlist_sheet)

# 9단계: 유통기한 업데이트
update_expiration(scraped_data, itemlist_barcodes, itemlist_sheet)

# 10단계: Status 열 업데이트
update_status_column(itemlist_sheet)

# 11단계: Variant SKU 업데이트
update_variant_sku(itemlist_sheet)

# 12단계: 재고가 0인 경우 가격 조정
adjust_prices_for_zero_stock(itemlist_sheet)

# 13단계: 업데이트된 Itemlist 파일 저장
if itemlist_sheet:
    updated_itemlist_file = f'Updated_Itemlist_{current_date}.xlsx'
    try:
        itemlist_wb.save(updated_itemlist_file)
        print(f"Updated Itemlist saved as '{updated_itemlist_file}'.")
    except Exception as e:
        print(f"Error saving updated itemlist file: {e}")
