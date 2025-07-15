import openpyxl
import csv
import re
import os
from datetime import datetime

# 현재 날짜 가져오기 (yymmdd 형식)
current_date = datetime.now().strftime('%y%m%d')

# 파일 경로 설정
scraped_file = f'Scraped_{current_date}.xlsx'
itemlist_file = f'Itemlist_{current_date}.xlsx'
scraped_m_file = f'ScrapedM_{current_date}.xlsx'
purchase_file = f'Purchase_{current_date}.xlsx'  # 추가된 Purchase 파일

print(f"Loading scraped file: {scraped_file}")
try:
    scraped_wb = openpyxl.load_workbook(scraped_file, read_only=False)
    scraped_sheet = scraped_wb.active
    print("Scraped file loaded successfully.")
except Exception as e:
    print(f"Error loading scraped file: {e}")
    scraped_sheet = None

print(f"Loading scrapedM file: {scraped_m_file}")
try:
    scraped_m_wb = openpyxl.load_workbook(scraped_m_file, read_only=False)
    scraped_m_sheet = scraped_m_wb.active
    print("ScrapedM file loaded successfully.")
except Exception as e:
    print(f"Error loading scrapedM file: {e}")
    scraped_m_sheet = None

print(f"Loading itemlist file: {itemlist_file}")
try:
    itemlist_wb = openpyxl.load_workbook(itemlist_file, read_only=False)
    itemlist_sheet = itemlist_wb.active
    print("Itemlist file loaded successfully.")
except Exception as e:
    print(f"Error loading itemlist file: {e}")
    itemlist_sheet = None

scraped_data = {}

if scraped_sheet:
    print("Reading scraped data...")
    try:
        for row in scraped_sheet.iter_rows(min_row=2, values_only=True):
            barcode = str(row[0])
            title = row[1]

            cr_stock = row[2] if row[2] is not None else 0
            cr_discount = row[3] if row[3] is not None else None
            cr_expiration = row[4] if row[4] else None

            nb_stock = row[5] if row[5] is not None else 0
            nb_discount = row[6] if row[6] is not None else None
            nb_expiration = row[7] if row[7] else None

            in_stock = row[8] if row[8] is not None else 0
            in_discount = row[9] if row[9] is not None else None
            in_expiration = row[10] if row[10] else None

            br_stock = row[11] if row[11] is not None else 0
            br_discount = row[12] if row[12] is not None else None
            br_expiration = row[13] if row[13] else None

            total_stock = cr_stock + nb_stock + in_stock + br_stock

            store_expiration_infos = {
                'CR': [(cr_expiration, cr_stock, cr_discount)],
                'NB': [(nb_expiration, nb_stock, nb_discount)],
                'IN': [(in_expiration, in_stock, in_discount)],
                'BR': [(br_expiration, br_stock, br_discount)]
            }

            expiration_infos = []
            for store_code in ['CR', 'NB', 'IN', 'BR']:
                expiration_infos.extend(store_expiration_infos[store_code])

            if expiration_infos:
                expiration_infos.sort(key=lambda x: x[0] if x[0] else '9999-12-31')
                processed_dates = set()
                found_valid_stock = False
                for exp_date, stock, discount in expiration_infos:
                    if exp_date in processed_dates:
                        continue
                    processed_dates.add(exp_date)
                    exp_total_stock = sum(s for ed, s, d in expiration_infos if ed == exp_date)
                    if exp_total_stock > 0:
                        total_stock = exp_total_stock
                        expiration_date = exp_date
                        discounts = [d for ed, s, d in expiration_infos if ed == exp_date and d is not None]
                        discount_price = discounts[0] if discounts else None
                        found_valid_stock = True
                        break
                if not found_valid_stock:
                    total_stock = 0
                    expiration_date = None
                    discount_price = None
            else:
                expiration_date = None
                discount_price = None

            store_stocks = {'CR': 0, 'NB': 0, 'IN': 0, 'BR': 0}
            if discount_price is not None:
                for store_code in ['CR', 'NB', 'IN', 'BR']:
                    store_exp_infos = store_expiration_infos[store_code]
                    store_stock_at_exp = sum(s for ed, s, d in store_exp_infos if ed == expiration_date)
                    store_stocks[store_code] = store_stock_at_exp
            else:
                store_stocks['CR'] = cr_stock
                store_stocks['NB'] = nb_stock
                store_stocks['IN'] = in_stock
                store_stocks['BR'] = br_stock

            scraped_data[barcode] = {
                'title': title,
                'total_stock': total_stock,
                'discount_price': discount_price,
                'expiration_date': expiration_date,
                'store_stocks': store_stocks
            }
        print("Scraped data read successfully.")
    except Exception as e:
        print(f"Error reading scraped data: {e}")

def read_scraped_m_data(scraped_m_sheet):
    scraped_m_data = {}
    for row in scraped_m_sheet.iter_rows(min_row=2, values_only=True):
        barcode = str(row[1])
        wholesaler_price = row[2]
        retail_price = row[3]
        scraped_m_data[barcode] = {
            'wholesaler_price': wholesaler_price,
            'retail_price': retail_price
        }
    return scraped_m_data

def get_itemlist_barcodes(itemlist_sheet):
    itemlist_barcodes = {}
    for row_idx in range(2, itemlist_sheet.max_row + 1):
        barcode = str(itemlist_sheet.cell(row=row_idx, column=14).value)
        itemlist_barcodes[barcode] = row_idx
    return itemlist_barcodes

def check_duplicate_barcodes(itemlist_sheet):
    barcodes_seen = set()
    duplicate_found = False
    for row_idx in range(2, itemlist_sheet.max_row + 1):
        barcode = str(itemlist_sheet.cell(row=row_idx, column=14).value)
        if barcode in barcodes_seen:
            print(f"Duplicate barcode found: {barcode} at row {row_idx}")
            duplicate_found = True
            exit(1)
        barcodes_seen.add(barcode)
    
    if not duplicate_found:
        print("중복된 바코드가 없습니다.")

def save_active_rows_to_csv(itemlist_sheet, output_csv_file):
    print("Saving active rows to CSV file based on Q and R columns...")
    try:
        headers = [cell.value for cell in itemlist_sheet[1]]
        active_rows = []
        for row_idx, row in enumerate(itemlist_sheet.iter_rows(min_row=2, values_only=True), start=2):
            q_value = str(row[16]).strip() if row[16] is not None else ""
            r_value = str(row[17]).strip() if row[17] is not None else ""
            if q_value == "O" and r_value == "O":
                row = list(row)
                if row[13] is not None:
                    row[13] = f'="{str(row[13])}"'
                for col_idx in [10, 11, 14]:
                    if row[col_idx] is not None and isinstance(row[col_idx], (float, int)):
                        row[col_idx] = f"{row[col_idx]:.2f}"
                row[15] = "active"
                active_rows.append(row)

        with open(output_csv_file, mode='w', newline='', encoding='utf-8-sig') as csvfile:
            writer = csv.writer(csvfile, quoting=csv.QUOTE_MINIMAL)
            writer.writerow(headers)
            writer.writerows(active_rows)

        print(f"Active rows saved successfully to '{output_csv_file}'. Rows saved: {len(active_rows)}")
    except Exception as e:
        print(f"Error saving active rows to CSV: {e}")

def compare_barcodes(scraped_data, itemlist_barcodes):
    scraped_barcodes = set(scraped_data.keys())
    itemlist_barcodes_set = set(itemlist_barcodes.keys())
    
    barcodes_to_add = scraped_barcodes - itemlist_barcodes_set
    barcodes_to_delete = itemlist_barcodes_set - scraped_barcodes
    return barcodes_to_add, barcodes_to_delete

def delete_barcodes(itemlist_barcodes, barcodes_to_delete, itemlist_sheet):
    rows_to_delete = [itemlist_barcodes[barcode] for barcode in barcodes_to_delete]
    rows_to_delete.sort(reverse=True)
    for row_idx in rows_to_delete:
        barcode = str(itemlist_sheet.cell(row=row_idx, column=14).value)
        print(f"Deleting row {row_idx} with barcode {barcode}")
        itemlist_sheet.delete_rows(row_idx)

def add_barcodes(barcodes_to_add, scraped_data, itemlist_sheet):
    for barcode in barcodes_to_add:
        new_row = [
            scraped_data[barcode]['title'],
            scraped_data[barcode]['title'],
            None,
            'TRUE',
            None,
            None,
            'shopify',
            scraped_data[barcode]['total_stock'],
            'deny',
            'manual',
            None,
            None,
            'TRUE',
            str(barcode),
            None,
            None,
            None,
            None,
            None,
            None
        ]
        last_row = itemlist_sheet.max_row + 1
        print(f"Adding new barcode {barcode} at row {last_row}")
        for col_idx, value in enumerate(new_row, start=1):
            cell = itemlist_sheet.cell(row=last_row, column=col_idx, value=value)
            if col_idx in [11, 12, 15]:
                if value is not None:
                    cell.number_format = '0.00'

def reorder_barcodes(scraped_data, itemlist_sheet):
    scraped_barcodes = list(scraped_data.keys())
    itemlist_barcodes = get_itemlist_barcodes(itemlist_sheet)
    itemlist_rows = [(barcode, row_idx) for barcode, row_idx in itemlist_barcodes.items()]
    new_order = []
    for barcode in scraped_barcodes:
        if barcode in itemlist_barcodes:
            row_idx = itemlist_barcodes[barcode]
            row_data = [cell.value for cell in itemlist_sheet[row_idx]]
            new_order.append(row_data)
        else:
            print(f"Barcode {barcode} not found in itemlist during reordering.")

    itemlist_sheet.delete_rows(2, itemlist_sheet.max_row - 1)
    for row_data in new_order:
        itemlist_sheet.append(row_data)
    print("Reordering completed.")

def update_prices(scraped_m_data, itemlist_sheet):
    itemlist_barcodes = get_itemlist_barcodes(itemlist_sheet)
    for barcode, row_idx in itemlist_barcodes.items():
        if barcode in scraped_m_data:
            data = scraped_m_data[barcode]
            k_cell = itemlist_sheet.cell(row=row_idx, column=11)
            l_cell = itemlist_sheet.cell(row=row_idx, column=12)
            o_cell = itemlist_sheet.cell(row=row_idx, column=15)

            k_cell.value = data['retail_price']
            l_cell.value = data['retail_price']
            k_cell.number_format = '0.00'
            l_cell.number_format = '0.00'

            # wholesaler_price가 0이 아닐 때만 O열에 값 작성
            if data['wholesaler_price'] != 0:
                o_cell.value = data['wholesaler_price']
                o_cell.number_format = '0.00'
            # 0이면 O열에 아무것도 하지 않음 (기존 값 유지)
        else:
            # ScrapedM에 해당 바코드가 없을 경우
            itemlist_sheet.cell(row=row_idx, column=11).value = 0
            itemlist_sheet.cell(row=row_idx, column=12).value = 0
            itemlist_sheet.cell(row=row_idx, column=15).value = 0
            print(f"Barcode {barcode} not found in ScrapedM. Prices set to 0.")

def update_existing_rows(scraped_data, itemlist_sheet):
    print("Matching and updating itemlist data...")
    try:
        for row_idx in range(2, itemlist_sheet.max_row + 1):
            variant_barcode = str(itemlist_sheet.cell(row=row_idx, column=14).value)
            if variant_barcode in scraped_data:
                row = itemlist_sheet[row_idx]
                if row[1].value != scraped_data[variant_barcode]['title']:
                    row[1].value = scraped_data[variant_barcode]['title']
                row[7].value = scraped_data[variant_barcode]['total_stock']
    except Exception as e:
        print(f"Error updating itemlist data: {e}")

def update_discount(scraped_data, itemlist_barcodes, itemlist_sheet):
    for barcode, data in scraped_data.items():
        if barcode in itemlist_barcodes:
            row_idx = itemlist_barcodes[barcode]
            if 'discount_price' in data and data['discount_price'] is not None:
                itemlist_sheet.cell(row=row_idx, column=11).value = data['discount_price']

def update_expiration(scraped_data, itemlist_barcodes, itemlist_sheet):
    for barcode, data in scraped_data.items():
        if barcode in itemlist_barcodes:
            row_idx = itemlist_barcodes[barcode]
            expiration_date = data['expiration_date']
            if expiration_date:
                expiration_text = f"Expiration date: {expiration_date}"
                itemlist_sheet.cell(row=row_idx, column=20).value = expiration_text
            else:
                itemlist_sheet.cell(row=row_idx, column=20).value = None

def update_store_stock_status(scraped_data, itemlist_barcodes, itemlist_sheet):
    headers = {
        21: 'Stock Status CR (product.metafields.custom.stock_cr)',
        22: 'Stock Status NB (product.metafields.custom.stock_nb)',
        23: 'Stock Status IN (product.metafields.custom.stock_in)',
        24: 'Stock Status BR (product.metafields.custom.stock_br)'
    }
    for col_idx, header in headers.items():
        itemlist_sheet.cell(row=1, column=col_idx).value = header

    for barcode, data in scraped_data.items():
        if barcode in itemlist_barcodes:
            row_idx = itemlist_barcodes[barcode]
            store_stocks = data['store_stocks']
            store_codes = ['CR', 'NB', 'IN', 'BR']
            for idx, store_code in enumerate(store_codes, start=21):
                stock = store_stocks[store_code]
                if stock == 0:
                    status = 'Out of stock'
                elif stock >= 10:
                    status = 'In stock'
                else:
                    status = f'{stock} Left'
                itemlist_sheet.cell(row=row_idx, column=idx).value = status

def update_variant_sku(itemlist_sheet):
    for row_idx in range(2, itemlist_sheet.max_row + 1):
        value = itemlist_sheet.cell(row=row_idx, column=14).value  # N열의 값 읽기
        itemlist_sheet.cell(row=row_idx, column=5).value = value    # E열에 값 직접 할당

def update_status_column(itemlist_sheet):
    print("Updating Status column (P열)...")
    try:
        for row_idx in range(2, itemlist_sheet.max_row + 1):
            status_formula = f'=IF(AND(Q{row_idx}="O", R{row_idx}="O"), "active", "archived")'
            itemlist_sheet.cell(row=row_idx, column=16).value = status_formula
    except Exception as e:
        print(f"Error updating Status column: {e}")

def adjust_prices_for_zero_stock(itemlist_sheet):
    print("Adjusting prices for items with zero stock...")
    try:
        for row_idx in range(2, itemlist_sheet.max_row + 1):
            inventory_qty = itemlist_sheet.cell(row=row_idx, column=8).value
            if inventory_qty == 0 or inventory_qty == '0':
                variant_compare_at_price = itemlist_sheet.cell(row=row_idx, column=12).value
                itemlist_sheet.cell(row=row_idx, column=11).value = variant_compare_at_price
    except Exception as e:
        print(f"Error adjusting prices: {e}")

# Purchase 파일에서 Description(A열), Price(C열), Comment(E열)을 읽어들이는 함수
def read_purchase_data(purchase_sheet):
    purchase_data = {}  
    for row in purchase_sheet.iter_rows(min_row=2, values_only=True):
        description = str(row[0])  # A열: Description
        price = row[2]  # C열: Price
        comment = row[4] if len(row) > 4 else None  # E열: Comment
        purchase_data[description] = {
            'price': price,
            'comment': comment
        }
    return purchase_data

# Purchase 정보로 Itemlist 업데이트 (O열에 Price, Y열(25열)에 Comment)
def update_cost_per_item_from_purchase(purchase_data, itemlist_sheet):
    for row_idx in range(2, itemlist_sheet.max_row + 1):
        original_title = str(itemlist_sheet.cell(row=row_idx, column=2).value)  # B열 Title
        # 괄호와 괄호 안의 내용을 제거한 뒤 앞뒤 공백 정리
        normalized_title = re.sub(r"\(.*?\)|\[.*?\]", "", original_title).strip()
        
        if normalized_title in purchase_data:
            cost_cell = itemlist_sheet.cell(row=row_idx, column=15)  # O열: Cost per item
            wholesaler_cell = itemlist_sheet.cell(row=row_idx, column=25) # Y열: Wholesesaler

            cost_value = purchase_data[normalized_title]['price']
            comment_value = purchase_data[normalized_title]['comment']

            cost_cell.value = cost_value
            cost_cell.number_format = '0.00'

            # Comment 정보 Y열에 작성
            if comment_value is not None:
                wholesaler_cell.value = comment_value
            else:
                wholesaler_cell.value = None
        else:
            # Purchase Data에 없는 경우 처리 (필요하다면)
            pass

def update_special_tag(itemlist_sheet):
    """
    1) Variant Price(K열, col=11)와 Variant Compare At Price(L열, col=12)를 비교
       - 같다면 C열에 ' Special'이라는 텍스트가 없어야 함(이미 있으면 제거)
       - 다르면 C열에 ' Special'이 있어야 함(없으면 추가)
    """
    for row_idx in range(2, itemlist_sheet.max_row + 1):
        # 각각의 셀 값 가져오기
        variant_price = itemlist_sheet.cell(row=row_idx, column=11).value
        compare_price = itemlist_sheet.cell(row=row_idx, column=12).value
        tags_cell = itemlist_sheet.cell(row=row_idx, column=3)

        # Tags 셀이 None인 경우 빈 문자열로 처리
        if tags_cell.value is None:
            tags_cell.value = ""

        # 문자열 형태로 변환
        tags_str = str(tags_cell.value)
        special_text = ", Special"

        # (1) 가격이 같으면 ', Special' 제거
        if variant_price == compare_price:
            if special_text in tags_str:
                new_tags_str = tags_str.replace(special_text, "")
                new_tags_str = new_tags_str.strip()  # 앞뒤 공백 정리
                tags_cell.value = new_tags_str
        else:
            # (2) 가격이 다르면 ', Special' 추가
            if special_text not in tags_str:
                # 기존 태그가 비어있지 않다면, 앞에 공백 포함해서 붙일지 여부 결정
                # 여기서는 바로 이어붙이는 방식으로 예시
                new_tags_str = tags_str.strip() + special_text
                tags_cell.value = new_tags_str

# 실행 흐름
check_duplicate_barcodes(itemlist_sheet)
itemlist_barcodes = get_itemlist_barcodes(itemlist_sheet)
barcodes_to_add, barcodes_to_delete = compare_barcodes(scraped_data, itemlist_barcodes)

delete_barcodes(itemlist_barcodes, barcodes_to_delete, itemlist_sheet)
add_barcodes(barcodes_to_add, scraped_data, itemlist_sheet)
check_duplicate_barcodes(itemlist_sheet)

itemlist_barcodes = get_itemlist_barcodes(itemlist_sheet)
reorder_barcodes(scraped_data, itemlist_sheet)
update_existing_rows(scraped_data, itemlist_sheet)

scraped_m_data = read_scraped_m_data(scraped_m_sheet)
update_prices(scraped_m_data, itemlist_sheet)

itemlist_barcodes = get_itemlist_barcodes(itemlist_sheet)
update_discount(scraped_data, itemlist_barcodes, itemlist_sheet)

print(f"Loading purchase file: {purchase_file}")
try:
    purchase_wb = openpyxl.load_workbook(purchase_file, read_only=False)
    purchase_sheet = purchase_wb.active
    print("Purchase file loaded successfully.")
    purchase_data = read_purchase_data(purchase_sheet)
    # Purchase 데이터 반영 (O열에 Price, Y열(25열)에 Comment)
    update_cost_per_item_from_purchase(purchase_data, itemlist_sheet)
except Exception as e:
    print(f"Error loading or processing purchase file: {e}")

itemlist_barcodes = get_itemlist_barcodes(itemlist_sheet)
update_expiration(scraped_data, itemlist_barcodes, itemlist_sheet)
update_store_stock_status(scraped_data, itemlist_barcodes, itemlist_sheet)
update_status_column(itemlist_sheet)
update_variant_sku(itemlist_sheet)
adjust_prices_for_zero_stock(itemlist_sheet)
update_special_tag(itemlist_sheet)

if itemlist_sheet:
    # 1) 임시 파일 이름 정의 (Itemlist_yymmdd_temp.xlsx)
    temp_file = itemlist_file.replace(".xlsx", "_temp.xlsx")

    try:
        # 2) 임시 파일로 먼저 저장
        itemlist_wb.save(temp_file)
        print(f"임시 파일로 저장 완료: {temp_file}")

        # 3) 기존 Itemlist_yymmdd.xlsx 파일이 있으면 삭제
        if os.path.exists(itemlist_file):
            os.remove(itemlist_file)
            print(f"기존 파일 삭제 완료: {itemlist_file}")

        # 4) 임시 파일을 최종 파일 이름으로 변경
        os.rename(temp_file, itemlist_file)
        print(f"임시 파일을 최종 이름으로 변경: {itemlist_file}")

    except Exception as e:
        print(f"Error saving or renaming itemlist file: {e}")

    # CSV 저장 로직(기존대로)
    active_csv_file = f'Active_Items_{current_date}.csv'
    save_active_rows_to_csv(itemlist_sheet, active_csv_file)
