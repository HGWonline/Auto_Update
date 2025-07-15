import openpyxl
import csv
from datetime import datetime

# 현재 날짜 (yymmdd)
current_date = datetime.now().strftime('%y%m%d')

# 파일명 설정
itemlist_file = f'Itemlist_{current_date}.xlsx'
promotion_file = f'Promotion_{current_date}.csv'

# Itemlist 엑셀 로드
try:
    itemlist_wb = openpyxl.load_workbook(itemlist_file, read_only=True, data_only=True)
    itemlist_sheet = itemlist_wb.active
except Exception as e:
    print(f"Error loading Itemlist file: {e}")
    itemlist_sheet = None

# Promotion 파일 헤더
headers = ['Name', 'Rate', 'Price', 'End date', 'Location', 'Online']
promotion_data = []

if itemlist_sheet:
    # 최대 3000행까지만 처리
    max_row_to_process = min(itemlist_sheet.max_row, 3000)
    rows = itemlist_sheet.iter_rows(min_row=2, max_row=max_row_to_process, values_only=True)

    for idx, row in enumerate(rows, start=2):
        # 인덱스: A(0), B(1), ... K(10), L(11), P(15), T(19), U(20), V(21), W(22), X(23)
        title = row[1]
        variant_price = row[10]
        variant_compare_price = row[11]
        status = row[15]
        expiration_cell = row[19]
        stock_cr = row[20]
        stock_nb = row[21]
        stock_in = row[22]
        stock_br = row[23]

        if variant_price is None or variant_compare_price is None:
            continue

        if variant_price != variant_compare_price:
            # Name
            name = title if title else ""

            # Price
            price = variant_price if variant_price else 0

            # Rate 계산: (l - k) / l * 100
            try:
                k = float(variant_price)
                l = float(variant_compare_price)
                if l != 0:
                    rate_val = (l - k) / l * 100
                    # 정수 여부와 관계없이 그대로 출력 (소수점 표현 유지)
                    rate_str = f"{rate_val}%"
                else:
                    rate_str = "0%"
            except:
                rate_str = "0%"

            # End date 추출
            end_date = ""
            if expiration_cell:
                str_val = str(expiration_cell)
                if "Expiration date:" in str_val:
                    end_date = str_val.replace("Expiration date:", "").strip()
                else:
                    end_date = str_val.strip()

            # Location
            store_mapping = [('CA', stock_cr), ('NB', stock_nb), ('IN', stock_in), ('BR', stock_br)]
            stores = [code for code, val in store_mapping if val is not None and "Out of stock" not in str(val)]
            if len(stores) == 4:
                location = "ALL"
            else:
                location = ", ".join(stores) if stores else ""

            # Online
            online = "O" if status == "active" else "X"

            promotion_data.append([name, rate_str, price, end_date, location, online])

        # 진행 상황 출력(옵션)
        if (idx % 500) == 0:
            print(f"Processed {idx} rows...")

# CSV 작성 시 utf-8-sig로 인코딩 변경
with open(promotion_file, 'w', newline='', encoding='utf-8-sig') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(headers)
    writer.writerows(promotion_data)

print(f"Promotion data saved to {promotion_file}")
