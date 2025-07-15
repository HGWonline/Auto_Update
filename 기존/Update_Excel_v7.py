import openpyxl
from datetime import datetime

# 현재 날짜 가져오기 (yymmdd 형식)
current_date = datetime.now().strftime('%y%m%d')

# 파일 경로 설정
scraped_file = f'Scraped_{current_date}.xlsx'
itemlist_file = f'Itemlist_{current_date}.xlsx'
scraped_m_file = f'ScrapedM_{current_date}.xlsx'  # 추가된 ScrapedM 파일

# Scraped_yymmdd.xlsx 로드 (시트 이름 지정)
print(f"Loading scraped file: {scraped_file}")
try:
    scraped_wb = openpyxl.load_workbook(scraped_file, read_only=False)
    scraped_sheet = scraped_wb.active  # 첫 번째 시트 사용
    print("Scraped file loaded successfully.")
except Exception as e:
    print(f"Error loading scraped file: {e}")
    scraped_sheet = None

# ScrapedM_yymmdd.xlsx 로드 (시트 이름 지정)
print(f"Loading scrapedM file: {scraped_m_file}")
try:
    scraped_m_wb = openpyxl.load_workbook(scraped_m_file, read_only=False)
    scraped_m_sheet = scraped_m_wb.active  # 첫 번째 시트 사용
    print("ScrapedM file loaded successfully.")
except Exception as e:
    print(f"Error loading scrapedM file: {e}")
    scraped_m_sheet = None

# Itemlist_yymmdd.xlsx 로드 (시트 이름 지정)
print(f"Loading itemlist file: {itemlist_file}")
try:
    itemlist_wb = openpyxl.load_workbook(itemlist_file, read_only=False)
    itemlist_sheet = itemlist_wb.active  # 첫 번째 시트 사용
    print("Itemlist file loaded successfully.")
except Exception as e:
    print(f"Error loading itemlist file: {e}")
    itemlist_sheet = None

# Scraped 데이터 저장을 위한 사전
scraped_data = {}

# Scraped 파일에서 데이터 읽기 (Barcode, Title, 총 재고, 할인 정보, 유통기한)
if scraped_sheet:
    print("Reading scraped data...")
    try:
        for row in scraped_sheet.iter_rows(min_row=2, values_only=True):
            barcode = str(row[0])  # Scraped에서 BarCode
            title = row[1]  # Description

            # 4개 매장의 재고 정보 가져오기
            cr_stock = row[2] if row[2] is not None else 0  # CR_Stock
            cr_discount = row[3] if row[3] is not None else None  # CR_Discount
            cr_expiration = row[4] if row[4] else None  # CR_Expiration

            nb_stock = row[5] if row[5] is not None else 0  # NB_Stock
            nb_discount = row[6] if row[6] is not None else None  # NB_Discount
            nb_expiration = row[7] if row[7] else None  # NB_Expiration

            in_stock = row[8] if row[8] is not None else 0  # IN_Stock
            in_discount = row[9] if row[9] is not None else None  # IN_Discount
            in_expiration = row[10] if row[10] else None  # IN_Expiration

            br_stock = row[11] if row[11] is not None else 0  # BR_Stock
            br_discount = row[12] if row[12] is not None else None  # BR_Discount
            br_expiration = row[13] if row[13] else None  # BR_Expiration

            # 총 재고 합산 (초기값은 모든 재고의 합)
            total_stock = cr_stock + nb_stock + in_stock + br_stock

            # 유통기한이 있는 매장의 재고 및 할인 정보 가져오기
            expiration_infos = []
            if cr_expiration:
                expiration_infos.append((cr_expiration, cr_stock, cr_discount))
            if nb_expiration:
                expiration_infos.append((nb_expiration, nb_stock, nb_discount))
            if in_expiration:
                expiration_infos.append((in_expiration, in_stock, in_discount))
            if br_expiration:
                expiration_infos.append((br_expiration, br_stock, br_discount))

            if expiration_infos:
                # 유통기한 순서대로 정렬
                expiration_infos.sort(key=lambda x: x[0])

                # 유통기한별로 재고 합산하여 재고가 0이 아닌 가장 빠른 유통기한 찾기
                processed_dates = set()
                found_valid_stock = False

                for exp_date, stock, discount in expiration_infos:
                    if exp_date in processed_dates:
                        continue
                    processed_dates.add(exp_date)

                    # 해당 유통기한의 재고 합산
                    exp_total_stock = sum(s for ed, s, d in expiration_infos if ed == exp_date)

                    if exp_total_stock > 0:
                        # 재고가 0보다 크면 해당 유통기한 사용
                        total_stock = exp_total_stock
                        expiration_date = exp_date
                        # 할인 가격 설정
                        discounts = [d for ed, s, d in expiration_infos if ed == exp_date and d is not None]
                        discount_price = discounts[0] if discounts else None
                        found_valid_stock = True
                        break  # 유효한 재고를 찾았으므로 반복 종료

                if not found_valid_stock:
                    # 모든 유통기한에 대해 재고가 0인 경우
                    total_stock = 0
                    expiration_date = None
                    discount_price = None
            else:
                # 유통기한이 없으면 None
                expiration_date = None
                discount_price = None

            scraped_data[barcode] = {
                'title': title,
                'total_stock': total_stock,
                'discount_price': discount_price,
                'expiration_date': expiration_date
            }
        print("Scraped data read successfully.")
    except Exception as e:
        print(f"Error reading scraped data: {e}")

# ScrapedM 데이터 읽기
def read_scraped_m_data(scraped_m_sheet):
    scraped_m_data = {}
    for row in scraped_m_sheet.iter_rows(min_row=2, values_only=True):
        barcode = str(row[1])  # ScrapedM에서 Barcode
        wholesaler_price = row[2]  # wholesalerPrice
        retail_price = row[3]  # retailPrice

        scraped_m_data[barcode] = {
            'wholesaler_price': wholesaler_price,
            'retail_price': retail_price
        }
    return scraped_m_data

# Itemlist 바코드 읽기
def get_itemlist_barcodes(itemlist_sheet):
    itemlist_barcodes = {}
    for row_idx in range(2, itemlist_sheet.max_row + 1):
        barcode = str(itemlist_sheet.cell(row=row_idx, column=14).value)  # N열의 Variant Barcode
        itemlist_barcodes[barcode] = row_idx
    return itemlist_barcodes

# 중복 바코드 확인 함수
def check_duplicate_barcodes(itemlist_sheet):
    barcodes_seen = set()
    duplicate_found = False
    for row_idx in range(2, itemlist_sheet.max_row + 1):
        barcode = str(itemlist_sheet.cell(row=row_idx, column=14).value)  # N열의 Variant Barcode
        if barcode in barcodes_seen:
            print(f"Duplicate barcode found: {barcode} at row {row_idx}")
            duplicate_found = True
            exit(1)  # 중복된 바코드가 있으면 작업 종료
        barcodes_seen.add(barcode)
    
    if not duplicate_found:
        print("중복된 바코드가 없습니다.")

# 1. 바코드를 비교하여 추가 및 삭제할 바코드 확인
def compare_barcodes(scraped_data, itemlist_barcodes):
    scraped_barcodes = set(scraped_data.keys())
    itemlist_barcodes_set = set(itemlist_barcodes.keys())
    
    # 추가해야 하는 바코드
    barcodes_to_add = scraped_barcodes - itemlist_barcodes_set
    
    # 삭제해야 하는 바코드
    barcodes_to_delete = itemlist_barcodes_set - scraped_barcodes
    
    return barcodes_to_add, barcodes_to_delete

# 2. 삭제할 바코드의 행을 삭제 (뒤에서부터 삭제)
def delete_barcodes(itemlist_barcodes, barcodes_to_delete, itemlist_sheet):
    rows_to_delete = [itemlist_barcodes[barcode] for barcode in barcodes_to_delete]
    rows_to_delete.sort(reverse=True)  # 역순으로 정렬하여 뒤에서부터 삭제
    for row_idx in rows_to_delete:
        barcode = str(itemlist_sheet.cell(row=row_idx, column=14).value)
        print(f"Deleting row {row_idx} with barcode {barcode}")
        itemlist_sheet.delete_rows(row_idx)

# 3. 추가할 바코드를 Itemlist의 마지막에 추가
def add_barcodes(barcodes_to_add, scraped_data, itemlist_sheet):
    for barcode in barcodes_to_add:
        new_row = [
            scraped_data[barcode]['title'],  # Handle (A열: Title과 동일)
            scraped_data[barcode]['title'],  # Title (B열)
            None,  # Tags (C열)
            'TRUE',  # Published (D열)
            None,  # Variant SKU (E열)
            None,  # Variant Grams (F열)
            'shopify',  # Variant Inventory Tracker (G열)
            scraped_data[barcode]['total_stock'],  # Variant Inventory Qty (H열)
            'deny',  # Variant Inventory Policy (I열)
            'manual',  # Variant Fulfillment Service (J열)
            None,  # Variant Price (K열)
            None,  # Variant Compare At Price (L열)
            'TRUE',  # Variant Taxable (M열)
            int(barcode),  # Variant Barcode (N열)
            None,  # Cost per item (O열)
            None,  # Status (P열)
            None,  # Gift Card (Q열)
            None,  # SEO Title (R열)
            None,  # SEO Description (S열)
            scraped_data[barcode]['expiration_date'],  # metafield.custom.expiration_date (T열, 20번째 열)
            None  # 마지막에 필요한 경우 추가
        ]
        last_row = itemlist_sheet.max_row + 1
        print(f"Adding new barcode {barcode} at row {last_row}")
        for col_idx, value in enumerate(new_row, start=1):
            itemlist_sheet.cell(row=last_row, column=col_idx, value=value)

# 4. 효율적으로 바코드를 재배열하는 함수
def reorder_barcodes(scraped_data, itemlist_sheet):
    """
    Scraped 파일의 바코드 순서에 맞게 Itemlist 파일의 바코드를 효율적으로 재배열.
    최소한의 이동만으로 최적화된 순서로 바코드를 재정렬.
    """
    # Scraped 데이터의 바코드 리스트
    scraped_barcodes = list(scraped_data.keys())
    
    # Itemlist의 바코드 리스트 (최신화된 상태)
    itemlist_barcodes = get_itemlist_barcodes(itemlist_sheet)
    
    # 바코드와 행 번호의 매핑을 리스트로 생성
    itemlist_rows = [(barcode, row_idx) for barcode, row_idx in itemlist_barcodes.items()]
    
    # Scraped 바코드 순서대로 Itemlist의 행을 재정렬
    new_order = []
    for barcode in scraped_barcodes:
        if barcode in itemlist_barcodes:
            row_idx = itemlist_barcodes[barcode]
            row_data = [cell.value for cell in itemlist_sheet[row_idx]]
            new_order.append(row_data)
        else:
            print(f"Barcode {barcode} not found in itemlist during reordering.")

    # 기존 데이터를 지우고, 헤더를 제외하고 재작성
    itemlist_sheet.delete_rows(2, itemlist_sheet.max_row - 1)
    for row_data in new_order:
        itemlist_sheet.append(row_data)

    print("Reordering completed.")

# 5. ScrapedM 파일의 retailPrice 및 wholesalerPrice를 Itemlist에 업데이트 (retailPrice는 K열과 L열에 작성)
def update_prices(scraped_m_data, itemlist_sheet):
    itemlist_barcodes = get_itemlist_barcodes(itemlist_sheet)  # 최신화된 바코드 맵
    for barcode, data in scraped_m_data.items():
        if barcode in itemlist_barcodes:
            row_idx = itemlist_barcodes[barcode]
            # retailPrice를 K열과 L열에 업데이트
            itemlist_sheet.cell(row=row_idx, column=11).value = data['retail_price']  # K열
            itemlist_sheet.cell(row=row_idx, column=12).value = data['retail_price']  # L열
            # wholesalerPrice를 O열에 업데이트
            itemlist_sheet.cell(row=row_idx, column=15).value = data['wholesaler_price']  # O열
            print(f"Updated prices for barcode {barcode}: retailPrice={data['retail_price']}, wholesalerPrice={data['wholesaler_price']}")
        else:
            print(f"Barcode {barcode} not found in itemlist.")

# 6. Scraped 데이터를 기반으로 Description과 총 재고를 각각 Title과 Variant Inventory Qty로 매칭
def update_existing_rows(scraped_data, itemlist_sheet):
    print("Matching and updating itemlist data...")
    try:
        for row_idx in range(2, itemlist_sheet.max_row + 1):
            variant_barcode = str(itemlist_sheet.cell(row=row_idx, column=14).value)  # N열의 Variant Barcode
            if variant_barcode in scraped_data:
                row = itemlist_sheet[row_idx]

                # Title 업데이트 (B열)
                if row[1].value != scraped_data[variant_barcode]['title']:
                    row[1].value = scraped_data[variant_barcode]['title']

                # Variant Inventory Qty 업데이트 (H열, 인덱스는 8)
                row[7].value = scraped_data[variant_barcode]['total_stock']
    except Exception as e:
        print(f"Error updating itemlist data: {e}")

# 7. Scraped 데이터를 기반으로 discount_price를 K열에 업데이트 (정보가 없으면 기존 retailPrice 유지)
def update_discount(scraped_data, itemlist_barcodes, itemlist_sheet):
    for barcode, data in scraped_data.items():
        if barcode in itemlist_barcodes:
            row_idx = itemlist_barcodes[barcode]
            # discount_price가 있는 경우에만 K열에 덮어쓰기
            if 'discount_price' in data and data['discount_price'] is not None:
                itemlist_sheet.cell(row=row_idx, column=11).value = data['discount_price']  # K열

# 8단계: Scraped 데이터를 기반으로 expiration_date를 metafield.custom.expiration_date (T열)에 업데이트
def update_expiration(scraped_data, itemlist_barcodes, itemlist_sheet):
    for barcode, data in scraped_data.items():
        if barcode in itemlist_barcodes:
            row_idx = itemlist_barcodes[barcode]
            expiration_date = data['expiration_date']
            
            # Expiration Date가 있는 경우
            if expiration_date:
                expiration_text = f"Expiration date: {expiration_date}"
                itemlist_sheet.cell(row=row_idx, column=20).value = expiration_text  # T열은 20번째 열
            else:
                itemlist_sheet.cell(row=row_idx, column=20).value = None  # 정보가 없으면 빈칸 유지

# Variant SKU 업데이트 (E열)
def update_variant_sku(itemlist_sheet):
    for row_idx in range(2, itemlist_sheet.max_row + 1):
        formula = f"=N{row_idx}"  # N열의 Barcode와 같게
        itemlist_sheet.cell(row=row_idx, column=5).value = formula  # E열에 수식 입력

# Status 열 (P열) 업데이트
def update_status_column(itemlist_sheet):
    print("Updating Status column (P열)...")
    try:
        for row_idx in range(2, itemlist_sheet.max_row + 1):
            status_formula = f'=IF(AND(Q{row_idx}="O", R{row_idx}="O"), "active", "archived")'
            itemlist_sheet.cell(row=row_idx, column=16).value = status_formula  # P열에 수식 입력 (인덱스는 16)
    except Exception as e:
        print(f"Error updating Status column: {e}")

# **추가된 함수**
# 12단계: 재고의 총합이 0인 경우 Variant Price(K열)을 Variant Compare At Price(L열)로 설정
def adjust_prices_for_zero_stock(itemlist_sheet):
    print("Adjusting prices for items with zero stock...")
    try:
        for row_idx in range(2, itemlist_sheet.max_row + 1):
            inventory_qty = itemlist_sheet.cell(row=row_idx, column=8).value  # H열: Variant Inventory Qty
            if inventory_qty == 0 or inventory_qty == '0':
                variant_compare_at_price = itemlist_sheet.cell(row=row_idx, column=12).value  # L열: Variant Compare At Price
                itemlist_sheet.cell(row=row_idx, column=11).value = variant_compare_at_price  # K열: Variant Price를 L열과 동일하게 설정
    except Exception as e:
        print(f"Error adjusting prices: {e}")

# 코드 실행
# 1단계: Itemlist 파일 로드 후 중복 바코드 검사
check_duplicate_barcodes(itemlist_sheet)

# 2단계: 바코드를 비교하여 추가 및 삭제할 바코드 확인
itemlist_barcodes = get_itemlist_barcodes(itemlist_sheet)
barcodes_to_add, barcodes_to_delete = compare_barcodes(scraped_data, itemlist_barcodes)

# 3단계: 삭제할 바코드를 제거
delete_barcodes(itemlist_barcodes, barcodes_to_delete, itemlist_sheet)

# 4단계: 추가할 바코드를 가장 아래에 추가 + 바코드 중복 검사
add_barcodes(barcodes_to_add, scraped_data, itemlist_sheet)
check_duplicate_barcodes(itemlist_sheet)

# 5단계: 바코드 순서 정렬
itemlist_barcodes = get_itemlist_barcodes(itemlist_sheet)
reorder_barcodes(scraped_data, itemlist_sheet)

# 6단계: Description과 총 재고를 업데이트
update_existing_rows(scraped_data, itemlist_sheet)

# 7단계: ScrapedM 데이터를 기반으로 retailPrice와 wholesalerPrice를 업데이트
scraped_m_data = read_scraped_m_data(scraped_m_sheet)
update_prices(scraped_m_data, itemlist_sheet)

# 8단계: discount_price를 K열에 업데이트 (정보가 없으면 기존 retailPrice 유지)
itemlist_barcodes = get_itemlist_barcodes(itemlist_sheet)
update_discount(scraped_data, itemlist_barcodes, itemlist_sheet)

# 9단계: expiration_date를 metafield.custom.expiration_date (T열)에 업데이트
itemlist_barcodes = get_itemlist_barcodes(itemlist_sheet)
update_expiration(scraped_data, itemlist_barcodes, itemlist_sheet)

# 10단계: Status 열 업데이트 (P열)
update_status_column(itemlist_sheet)

# 11단계: Variant SKU 업데이트
update_variant_sku(itemlist_sheet)

# **12단계: 재고가 0인 경우 가격 조정**
adjust_prices_for_zero_stock(itemlist_sheet)

# 13단계: 업데이트된 Itemlist 파일 저장
if itemlist_sheet:
    updated_itemlist_file = f'Updated_Itemlist_{current_date}.xlsx'
    try:
        itemlist_wb.save(updated_itemlist_file)
        print(f"Updated Itemlist saved as '{updated_itemlist_file}'.")
    except Exception as e:
        print(f"Error saving updated itemlist file: {e}")
