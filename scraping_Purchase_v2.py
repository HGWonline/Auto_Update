import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime, timedelta
import openpyxl
import traceback
import os
import requests

def send_slack_message(text):
    webhook_url = "https://hooks.slack.com/services/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"
    payload = {"text": text}
    try:
        response = requests.post(webhook_url, json=payload)
        if response.status_code != 200:
            print(f"❗ 슬랙 알림 실패: {response.status_code}")
    except Exception as e:
        print(f"❗ 슬랙 알림 오류: {e}")

# Chrome 설정
chrome_service = Service(ChromeDriverManager().install())
chrome_options = Options()
chrome_options.add_argument('--headless')
chrome_options.add_argument('--disable-gpu')
chrome_options.add_argument('--window-size=1920,1080')

driver = webdriver.Chrome(service=chrome_service, options=chrome_options)

def wait_for_page_load(timeout=10):
    WebDriverWait(driver, timeout).until(
        lambda d: d.execute_script('return document.readyState') == 'complete'
    )

def login_and_navigate(driver):
    print("Navigating to login page...")
    driver.get('https://www.hangawee.com.au/login')

    username_input = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, 'username'))
    )
    print("Entering login code...")
    username_input.send_keys('245554073198')

    sign_in_button = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.ID, 'kt_login_signin_submit'))
    )
    print("Clicking sign in button...")
    sign_in_button.click()

    # Outlet 선택
    carousel_label = WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.XPATH, "//label[contains(., 'Carousel')]"))
    )
    print("Selecting Carousel outlet...")
    carousel_label.click()

    next_button = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.ID, 'next-step'))
    )
    print("Clicking Next button...")
    next_button.click()

    confirm_button = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.ID, 'goIndex'))
    )
    driver.execute_script("arguments[0].scrollIntoView(true);", confirm_button)
    time.sleep(1)
    print("Clicking Confirm button...")
    confirm_button.click()

    # 홈 페이지의 특정 요소가 로드될 때까지 대기
    print("Waiting for home page to fully load...")
    try:
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, 'sales-volume-chart'))
        )
    except:
        print("Home page chart not found, proceeding anyway.")

    # Purchase 페이지로 이동
    print("Navigating to purchase page…")
    driver.get("https://www.hangawee.com.au/?page=purchase")
    
    WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, 'button.btn.dropdown-toggle.btn-light'))
    )
    
def set_date_range_and_search(days_back=10):
    # days_back일 전 ~ 오늘 날짜
    start_date_obj = datetime.now() - timedelta(days=days_back)
    start_date = start_date_obj.strftime("%d/%m/%Y")
    end_date = datetime.now().strftime("%d/%m/%Y")

    start_date_input = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, 'startDate'))
    )
    end_date_input = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, 'endDate'))
    )

    start_date_input.clear()
    start_date_input.send_keys(start_date)
    end_date_input.clear()
    end_date_input.send_keys(end_date)

    # Purchase history 버튼 클릭
    purchase_history_button = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.ID, 'kt_search'))
    )
    purchase_history_button.click()

    wait_for_page_load()

def change_view_to_50():
    try:
        dropdown_toggle = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, ".datatable-pager-size .dropdown-toggle"))
        )
        dropdown_toggle.click()  # 드롭다운 열기

        view_50_element = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//li/a/span[text()='50']"))
        )
        view_50_element.click()

        wait_for_page_load()
        print("View changed to 50 successfully.")
    except Exception as e:
        print("Error changing view to 50:", e)
        traceback.print_exc()

def scrape_purchase_data():
    """
    새로 스크래핑한 데이터를 딕셔너리 형태로 반환:
    new_data = {
      description: (purchase_dt, price, user_name, comment)
    }
    """
    new_data = {}
    page = 1
    while True:
        print(f"Scraping page {page}...")
        # 테이블 로딩 대기
        try:
            rows = WebDriverWait(driver, 10).until(
                EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'table.datatable-table tbody.datatable-body tr'))
            )
        except:
            # 테이블 행이 없으면 종료
            print("No rows found, stopping.")
            break

        if not rows:
            print("No rows on this page, stopping.")
            break

        for row in rows:
            try:
                description = row.find_element(By.CSS_SELECTOR, 'td[data-field="description"]').text.strip()
                purchase_date = row.find_element(By.CSS_SELECTOR, 'td[data-field="purchaseDate"]').text.strip()
                price = row.find_element(By.CSS_SELECTOR, 'td[data-field="price"] input').get_attribute('value').strip()
                user_name = row.find_element(By.CSS_SELECTOR, 'td[data-field="realName"]').text.strip()
                
                # Comment 추출
                comment_element = row.find_element(By.CSS_SELECTOR, 'td[data-field="comment"] input')
                comment = comment_element.get_attribute('value').strip()

                # 날짜 파싱 (일/월/년)
                try:
                    purchase_dt = datetime.strptime(purchase_date, "%d/%m/%Y")
                except Exception as e:
                    print(f"Date parsing error for {purchase_date}: {e}")
                    purchase_dt = None

                # 동일 description에 대해 더 최신 날짜만 유지
                if description in new_data:
                    existing_dt = new_data[description][0]  # (purchase_dt, price, user_name, comment) 중 dt
                    if purchase_dt and existing_dt and purchase_dt > existing_dt:
                        new_data[description] = (purchase_dt, price, user_name, comment)
                else:
                    new_data[description] = (purchase_dt, price, user_name, comment)

            except Exception as e:
                print("Error scraping a row:", e)
                traceback.print_exc()
                continue

        # 다음 페이지 이동
        try:
            next_btn = driver.find_element(By.CSS_SELECTOR, 'a.datatable-pager-link[title="Next"]')
            next_btn_classes = next_btn.get_attribute('class')
            if 'datatable-pager-link-disabled' in next_btn_classes or next_btn.get_attribute('disabled') == 'disabled':
                print("Reached the last page.")
                break

            next_btn.click()
            time.sleep(3)
            wait_for_page_load()
            page += 1
        except:
            print("No next button found, stopping.")
            break

    return new_data


def load_existing_purchase_data(file_path):
    """
    기존 Purchase_xxxx.xlsx 파일을 열어,
    {description: (purchase_dt, price, user_name, comment)} 형태로 리턴.
    파일이 없거나 로드 실패 시 빈 dict 리턴.
    """
    if not os.path.exists(file_path):
        print(f"[load_existing_purchase_data] {file_path} does not exist, returning empty.")
        return {}
    
    existing_data = {}
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        
        # 시트 구조가 ['Description', 'Purchase Date', 'Price', 'User Name', 'Comment'] 라고 가정
        for row_idx in range(2, sheet.max_row + 1):
            desc_cell = sheet.cell(row=row_idx, column=1).value
            date_cell = sheet.cell(row=row_idx, column=2).value
            price_cell = sheet.cell(row=row_idx, column=3).value
            user_cell = sheet.cell(row=row_idx, column=4).value
            comment_cell = sheet.cell(row=row_idx, column=5).value
            
            if not desc_cell:
                continue
            
            description = str(desc_cell).strip()
            purchase_dt = None
            if date_cell:
                # 'dd/mm/yyyy' 형태로 들어있다고 가정
                try:
                    purchase_dt = datetime.strptime(str(date_cell), "%d/%m/%Y")
                except:
                    # 만약 엑셀 내 날짜가 datetime 객체라면 직접 변환
                    if isinstance(date_cell, datetime):
                        purchase_dt = date_cell
                    else:
                        purchase_dt = None

            price = str(price_cell).strip() if price_cell else ""
            user_name = str(user_cell).strip() if user_cell else ""
            comment = str(comment_cell).strip() if comment_cell else ""

            # 동일 description 있을 경우 -> 날짜 비교 후, 더 최신 데이터로 업데이트
            if description in existing_data:
                old_dt = existing_data[description][0]
                # old_dt가 None 이거나, 현재 dt가 더 최신이면 교체
                if not old_dt or (purchase_dt and purchase_dt > old_dt):
                    existing_data[description] = (purchase_dt, price, user_name, comment)
            else:
                existing_data[description] = (purchase_dt, price, user_name, comment)

        wb.close()
        print(f"Loaded existing data from {file_path}, total {len(existing_data)} items.")
        return existing_data
    except Exception as e:
        print(f"Error loading file {file_path}: {e}")
        traceback.print_exc()
        return {}


def merge_data(existing_data, new_data):
    """
    기존 데이터와 새 데이터를 병합하여 최종 dict 반환
    description을 키로 하여,
    - 새 데이터에 description이 있으면 날짜 비교 후 업데이트
    - 없으면 그대로 유지
    """
    merged_data = dict(existing_data)  # 먼저 기존 데이터 전체 복사
    
    for desc, (p_dt, p_price, p_user, p_comment) in new_data.items():
        if desc in merged_data:
            old_dt = merged_data[desc][0]
            if p_dt and old_dt and p_dt > old_dt:
                merged_data[desc] = (p_dt, p_price, p_user, p_comment)
            elif old_dt is None and p_dt:  # 기존이 None이면 새 데이터로 업데이트
                merged_data[desc] = (p_dt, p_price, p_user, p_comment)
            # 그 외 (날짜가 더 예전이거나 둘 다 None 등)은 기존 데이터 유지
        else:
            # 완전히 새로운 description
            merged_data[desc] = (p_dt, p_price, p_user, p_comment)
    
    return merged_data

def save_purchase_data_with_safe_rename(merged_data, old_file_path):
    """
    병합된 데이터(merged_data)를 임시 파일에 저장한 뒤,
    구(舊) 파일을 삭제하고 임시 파일 이름을 old_file_path로 변경
    """
    # 1) 임시 파일 이름 정의
    temp_file_path = old_file_path.replace(".xlsx", "_temp.xlsx")
    
    # 2) 병합된 데이터를 임시 파일로 저장
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Purchase Data"
    sheet.append(['Description', 'Purchase Date', 'Price', 'User Name', 'Comment'])

    for desc, (p_dt, p_price, p_user, p_comment) in merged_data.items():
        p_date_str = p_dt.strftime("%d/%m/%Y") if p_dt else ""
        sheet.append([desc, p_date_str, p_price, p_user, p_comment])

    wb.save(temp_file_path)
    wb.close()
    print(f"임시 파일로 저장 완료: {temp_file_path}")

    # 3) 기존 파일 제거
    if os.path.exists(old_file_path):
        try:
            os.remove(old_file_path)
            print(f"구(舊) 파일 삭제 완료: {old_file_path}")
        except Exception as e:
            print(f"구 파일을 삭제하지 못했습니다: {e}")

    # 4) 임시 파일을 최종 이름으로 변경
    try:
        os.rename(temp_file_path, old_file_path)
        print(f"임시 파일을 최종 이름으로 변경: {temp_file_path} -> {old_file_path}")
    except Exception as e:
        print(f"임시 파일을 {old_file_path}로 이름 변경 실패: {e}")

def save_backup_file(merged_data, backup_path):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.append(['Description', 'Purchase Date', 'Price', 'User Name', 'Comment'])
    for desc, (p_dt, p_price, p_user, p_comment) in merged_data.items():
        date_str = p_dt.strftime("%d/%m/%Y") if p_dt else ""
        sheet.append([desc, date_str, p_price, p_user, p_comment])
    wb.save(backup_path)
    wb.close()
    print(f"Backup saved to: {backup_path}")

def main():
    today_str = datetime.now().strftime("%y%m%d")

    # ✅ 누적 저장 및 백업 모두 오늘 날짜 기준 파일명
    file_name = f'Purchase_{today_str}.xlsx'

    # ✅ 저장 파일 (누적용)
    old_file_path = file_name

    # ✅ 백업 디렉토리 및 파일
    backup_dir = r'C:\Users\김남빈\OneDrive\★Hangaweemarket\Online\☆Item_List'
    os.makedirs(backup_dir, exist_ok=True)
    backup_path = os.path.join(backup_dir, file_name)

    try:
        existing_data = load_existing_purchase_data(old_file_path)

        login_and_navigate(driver)
        set_date_range_and_search(days_back=10)
        change_view_to_50()
        new_data = scrape_purchase_data()

        merged_data = merge_data(existing_data, new_data)

        # 백업 저장
        save_backup_file(merged_data, backup_path)

        # 누적 파일도 같은 이름으로 저장
        save_purchase_data_with_safe_rename(merged_data, old_file_path)
        send_slack_message(f"✅ {file_name}.xlsx 생성 완료!")

    except Exception as e:
        print("An unexpected error occurred during scraping:")
        traceback.print_exc()
        send_slack_message(f"❌ Purchase 크롤링 실패: {e}")

    finally:
        driver.quit()

# 이 부분이 반드시 있어야 main()이 실행됩니다!
if __name__ == "__main__":
    main()

