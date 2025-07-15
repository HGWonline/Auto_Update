import openpyxl
from datetime import datetime

# 현재 날짜 가져오기 (yymmdd 형식)
current_date = datetime.now().strftime('%y%m%d')

# 파일 경로 설정
itemlist_file = f'Itemlist_{current_date}.xlsx'
scraped_m_file = f'ScrapedM_{current_date}.xlsx'  # ScrapedM 파일

# ScrapedM_yymmdd.xlsx 로드 (시트 이름 지정)
print(f"Loading ScrapedM file: {scraped_m_file}")
try:
    scraped_m_wb = openpyxl.load_workbook(scraped_m_file, read_only=False)
    scraped_m_sheet = scraped_m_wb.active  # 첫 번째 시트 사용
    print("ScrapedM file loaded successfully.")
except Exception as e:
    print(f"Error loading ScrapedM file: {e}")
    scraped_m_sheet = None

# Itemlist_yymmdd.xlsx 로드 (시트 이름 지정)
print(f"Loading Itemlist file: {itemlist_file}")
try:
    itemlist_wb = openpyxl.load_workbook(itemlist_file, read_only=False)
    itemlist_sheet = itemlist_wb.active  # 첫 번째 시트 사용
    print("Itemlist file loaded successfully.")
except Exception as e:
    print(f"Error loading Itemlist file: {e}")
    itemlist_sheet = None

# ScrapedM 데이터 읽기
def read_scraped_m_data(scraped_m_sheet):
    scraped_m_data = {}
    for row in scraped_m_sheet.iter_rows(min_row=2, values_only=True):
        barcode = str(row[1])  # ScrapedM에서 Barcode (B열)
        wholesaler_price = row[2]  # wholesalerPrice (C열)
        retail_price = row[3]      # retailPrice (D열)

        scraped_m_data[barcode] = {
            'wholesaler_price': wholesaler_price,
            'retail_price': retail_price
        }
    return scraped_m_data

# Itemlist 바코드 읽기
def get_itemlist_barcodes(itemlist_sheet):
    itemlist_barcodes = {}
    row_idx = 2  # Start from the second row
    while row_idx <= itemlist_sheet.max_row:
        barcode_cell = itemlist_sheet.cell(row=row_idx, column=6)  # F열의 Variant Barcode
        barcode = barcode_cell.value
        if barcode is not None:
            barcode = str(barcode)
            # Collect the 4 rows corresponding to this barcode
            rows = [row_idx + i for i in range(4) if row_idx + i <= itemlist_sheet.max_row]
            itemlist_barcodes[barcode] = rows
            row_idx += 4  # Skip to the next set
        else:
            row_idx += 1  # Move to next row
    return itemlist_barcodes

# 중복 바코드 확인 함수
def check_duplicate_barcodes(itemlist_sheet):
    barcodes_seen = set()
    duplicate_found = False
    row_idx = 2
    while row_idx <= itemlist_sheet.max_row:
        barcode_cell = itemlist_sheet.cell(row=row_idx, column=6)  # F열의 Variant Barcode
        barcode = barcode_cell.value
        if barcode is not None:
            barcode = str(barcode)
            if barcode in barcodes_seen:
                print(f"Duplicate barcode found: {barcode} at row {row_idx}")
                duplicate_found = True
                exit(1)  # 중복된 바코드가 있으면 작업 종료
            barcodes_seen.add(barcode)
            row_idx += 4  # Skip to the next set
        else:
            row_idx += 1  # Move to next row

    if not duplicate_found:
        print("중복된 바코드가 없습니다.")

# ScrapedM 파일의 retailPrice 및 wholesalerPrice를 Itemlist에 업데이트
def update_prices(scraped_m_data, itemlist_sheet):
    itemlist_barcodes = get_itemlist_barcodes(itemlist_sheet)  # 최신화된 바코드 맵
    for barcode, data in scraped_m_data.items():
        if barcode in itemlist_barcodes:
            row_indices = itemlist_barcodes[barcode]
            for row_idx in row_indices:
                # I열과 J열의 값을 가져오기
                value_i = itemlist_sheet.cell(row=row_idx, column=9).value  # I열
                value_j = itemlist_sheet.cell(row=row_idx, column=10).value  # J열

                # I열과 J열의 정보가 동일한지 확인
                if value_i == value_j:
                    # retailPrice를 I열과 J열에 업데이트
                    itemlist_sheet.cell(row=row_idx, column=9).value = data['retail_price']   # I열
                    itemlist_sheet.cell(row=row_idx, column=10).value = data['retail_price']  # J열
                else:
                    # retailPrice를 J열에만 업데이트
                    itemlist_sheet.cell(row=row_idx, column=10).value = data['retail_price']  # J열

                # wholesalerPrice를 K열에 업데이트
                itemlist_sheet.cell(row=row_idx, column=11).value = data['wholesaler_price']  # K열

            print(f"Updated prices for barcode {barcode}: retailPrice={data['retail_price']}, wholesalerPrice={data['wholesaler_price']}")
        else:
            print(f"Barcode {barcode} not found in Itemlist.")

# 코드 실행
# 1단계: Itemlist 파일 로드 후 중복 바코드 검사
check_duplicate_barcodes(itemlist_sheet)

# 2단계: ScrapedM 데이터를 기반으로 retailPrice와 wholesalerPrice를 업데이트
scraped_m_data = read_scraped_m_data(scraped_m_sheet)
update_prices(scraped_m_data, itemlist_sheet)

# 3단계: 업데이트된 Itemlist 파일 저장
if itemlist_sheet:
    updated_itemlist_file = f'Updated_Itemlist_{current_date}.xlsx'
    try:
        itemlist_wb.save(updated_itemlist_file)
        print(f"Updated Itemlist saved as '{updated_itemlist_file}'.")
    except Exception as e:
        print(f"Error saving updated itemlist file: {e}")
