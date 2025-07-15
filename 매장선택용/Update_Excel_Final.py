import openpyxl
from datetime import datetime

# 현재 날짜 가져오기 (yymmdd 형식)
current_date = datetime.now().strftime('%y%m%d')

# 파일 경로 설정
scraped_file = f'Scraped_{current_date}.xlsx'
itemlist_file = f'Itemlist_{current_date}.xlsx'

# Scraped 파일 로드 (시트 이름 지정)
print(f"Loading scraped file: {scraped_file}")
try:
    scraped_wb = openpyxl.load_workbook(scraped_file, read_only=False)
    scraped_sheet = scraped_wb.active  # 첫 번째 시트 사용
    print("Scraped file loaded successfully.")
except Exception as e:
    print(f"Error loading scraped file: {e}")
    scraped_sheet = None

# Itemlist 파일 로드 (시트 이름 지정)
print(f"Loading itemlist file: {itemlist_file}")
try:
    itemlist_wb = openpyxl.load_workbook(itemlist_file, read_only=False)
    itemlist_sheet = itemlist_wb.active  # 첫 번째 시트 사용
    print("Itemlist file loaded successfully.")
except Exception as e:
    print(f"Error loading itemlist file: {e}")
    itemlist_sheet = None

# Scraped 데이터 저장을 위한 사전 (각 매장별 정보 저장)
scraped_data = {}

# Scraped 파일에서 데이터 읽기 (Barcode, Title, BR_Stock, BR_Discount, BR_Expiration 추출)
if scraped_sheet:
    print("Reading scraped data...")
    try:
        for row in scraped_sheet.iter_rows(min_row=2, values_only=True):
            barcode = int(row[0])  # Scraped에서 Barcode (첫 번째 열)
            title = row[1]  # Description (두 번째 열)
            ca_stock = row[2]  # CA_Stock
            ca_discount = row[3] if row[3] is not None else None # CA_Discount
            ca_expiration = row[4] if row[4] is not None else None # CA_Expiration
            nb_stock = row[5]  # NB_Stock
            nb_discount = row[6] if row[6] is not None else None # NB_Discount
            nb_expiration = row[7] if row[7] is not None else None # NB_Expiration
            in_stock = row[8]  # IN_Stock
            in_discount = row[9] if row[9] is not None else None # IN_Discount
            in_expiration = row[10] if row[10] is not None else None # IN_Expiration
            br_stock = row[11]  # BR_Stock
            br_discount = row[12] if row[12] is not None else None # BR_Discount
            br_expiration = row[13] if row[13] is not None else None # BR_Expiration

            # 정확한 데이터 분리를 위한 추가 처리
            scraped_data[barcode] = {
                'title': title,
                'ca_stock': ca_stock,
                'ca_discount': ca_discount,
                'ca_expiration': ca_expiration,
                'nb_stock': nb_stock,
                'nb_discount': nb_discount,
                'nb_expiration': nb_expiration,
                'in_stock': in_stock,
                'in_discount': in_discount,
                'in_expiration': in_expiration,
                'br_stock': br_stock,  # BR_Stock
                'br_discount': br_discount,  # BR_Discount
                'br_expiration': br_expiration  # BR_Expiration
            }
        print("Scraped data read successfully.")
    except Exception as e:
        print(f"Error reading scraped data: {e}")

# Itemlist 바코드 읽기 (4개 행씩 하나의 세트로 처리)
def get_itemlist_barcodes(itemlist_sheet):
    """
    Itemlist 파일에서 Booragoon 바코드(첫 번째 행)의 열을 참조하여 바코드를 추출.
    빈 값이나 None은 건너뛰도록 처리.
    """
    itemlist_barcodes = {}
    for row_idx in range(2, itemlist_sheet.max_row + 1, 4):  # 4개 행이 한 세트로 묶여 있음
        barcode_value = itemlist_sheet.cell(row=row_idx, column=6).value

        if barcode_value is None or barcode_value == "":
            continue  # 빈 값은 건너뛰기

        barcode = str(barcode_value)  # 문자열 형식으로 변환하여 저장
        itemlist_barcodes[barcode] = row_idx  # Booragoon 행의 인덱스를 저장

    return itemlist_barcodes

# 중복 바코드 확인 함수
def check_duplicate_barcodes(itemlist_sheet):
    """
    중복 바코드를 체크하는 과정에서 문자열 형식으로 바코드를 처리합니다.
    """
    barcodes_seen = set()
    duplicate_found = False

    for row_idx in range(2, itemlist_sheet.max_row + 1):
        barcode_value = itemlist_sheet.cell(row=row_idx, column=6).value

        if not barcode_value:  # 빈 값 처리
            continue
        
        barcode = str(barcode_value)  # 문자열로 변환

        if barcode in barcodes_seen:
            print(f"Duplicate barcode found: {barcode} at row {row_idx}")
            duplicate_found = True
            exit(1)  # 중복된 바코드가 있으면 작업 종료
        barcodes_seen.add(barcode)
    
    if not duplicate_found:
        print("중복된 바코드가 없습니다.")

# 1. 바코드를 비교하여 추가 및 삭제할 바코드 확인
def compare_barcodes(scraped_data, itemlist_barcodes):
    """
    Scraped 데이터와 Itemlist의 바코드를 비교하여 추가할 바코드와 삭제할 바코드를 반환.
    모든 바코드를 문자열 형식으로 통일하여 비교합니다.
    """
    scraped_barcodes = {str(barcode) for barcode in scraped_data.keys()}  # Scraped 바코드 문자열 변환
    itemlist_barcodes_set = {str(barcode) for barcode in itemlist_barcodes.keys()}  # Itemlist 바코드 문자열 변환
    
    # print(f"Scraped Barcodes: {scraped_barcodes}")
    # print(f"Itemlist Barcodes: {itemlist_barcodes_set}")
    
    barcodes_to_add = scraped_barcodes - itemlist_barcodes_set
    barcodes_to_delete = itemlist_barcodes_set - scraped_barcodes
    
    # print(f"Barcodes to add: {barcodes_to_add}")
    # print(f"Barcodes to delete: {barcodes_to_delete}")
    
    return barcodes_to_add, barcodes_to_delete

# 2. 삭제할 바코드의 행을 삭제 (같은 Handle 값의 4개 행을 삭제)
def delete_barcodes(itemlist_barcodes, barcodes_to_delete, itemlist_sheet):
    """
    삭제할 바코드의 행을 역순으로 삭제 (같은 Handle 값의 4개 행을 삭제)
    """
    barcodes_to_delete_sorted = sorted([itemlist_barcodes[barcode] for barcode in barcodes_to_delete], reverse=True)

    for row_idx in barcodes_to_delete_sorted:
        print(f"Deleting rows {row_idx} to {row_idx + 3}")
        itemlist_sheet.delete_rows(row_idx, amount=4)

# 3. 추가할 바코드를 Itemlist의 마지막에 추가 (한 번에 4개 행 추가)
def add_barcodes(barcodes_to_add, scraped_data, itemlist_sheet):
    for barcode in barcodes_to_add:
        location_values = ['Booragoon', 'Carousel', 'Northbridge', 'Innaloo']

        # 바코드를 추가할 마지막 행을 계산
        last_row = itemlist_sheet.max_row + 1
        print(f"Attempting to add barcode {barcode} at rows {last_row} to {last_row + 3}")

        for i, location in enumerate(location_values):
            # Booragoon인 경우에만 F열에 바코드 입력
            if location == 'Booragoon':
                try:
                    barcode_value = int(barcode)
                except ValueError:
                    print(f"Error converting barcode {barcode} to integer at row {last_row + i}")
                    continue  # 변환 실패 시 건너뛰기

            else:
                barcode_value = ''  # Booragoon이 아닌 행은 빈 값

            row = last_row + i

            # E열에 매장 이름, F열에 바코드를 입력
            itemlist_sheet.cell(row=row, column=5, value=location)  # E열: 매장 이름
            itemlist_sheet.cell(row=row, column=6, value=barcode_value)  # F열: Booragoon에만 바코드 기록

            # print(f"Written store '{location}' in E{row}, barcode '{barcode_value}' in F{row}")

    print(f"Successfully attempted to add {len(barcodes_to_add)} barcodes.")


# 4. 바코드 재정렬 최적화
def reorder_barcodes(scraped_data, itemlist_sheet):
    """
    Scraped 파일의 바코드 순서에 맞게 Itemlist 파일의 바코드와 관련된 모든 데이터를 재정렬.
    A열부터 Y열까지의 데이터를 포함한 모든 행을 메모리로 읽고, 재정렬한 후 다시 작성.
    """
    # Scraped 데이터의 바코드 리스트
    scraped_barcodes = list(scraped_data.keys())

    # Itemlist에서 바코드가 있는 첫 번째 행(Booragoon)의 인덱스를 추출
    itemlist_barcodes = get_itemlist_barcodes(itemlist_sheet)

    # 메모리 상에서 재정렬할 데이터를 저장할 리스트
    sorted_data = []

    # Scraped 데이터 순서에 맞게 Itemlist 데이터를 재정렬
    for barcode in scraped_barcodes:
        if str(barcode) in itemlist_barcodes:
            start_row = itemlist_barcodes[str(barcode)]
            rows_data = []
            for i in range(4):  # 4개의 행이 한 세트
                row_data = [cell.value for cell in itemlist_sheet[start_row + i][0:25]]  # A열부터 Y열까지 가져옴
                rows_data.append(row_data)

            sorted_data.extend(rows_data)
        else:
            print(f"Barcode {barcode} not found in Itemlist, skipping.")

    # 기존 Itemlist 시트의 모든 데이터를 삭제 (2번째 행부터 끝까지)
    itemlist_sheet.delete_rows(2, itemlist_sheet.max_row)

    # 재정렬된 데이터를 엑셀 시트에 다시 작성
    for row_idx, row_data in enumerate(sorted_data, start=2):
        for col_idx, cell_value in enumerate(row_data, start=1):
            itemlist_sheet.cell(row=row_idx, column=col_idx, value=cell_value)

    print("Reordering completed.")

# 매장 이름이 올바르게 입력되었는지 확인
def ensure_store_names(itemlist_sheet):
    store_names = ['Booragoon', 'Carousel', 'Northbridge', 'Innaloo']
    
    for row_idx in range(2, itemlist_sheet.max_row + 1, 4):
        for i, store_name in enumerate(store_names):
            cell_value = itemlist_sheet.cell(row=row_idx + i, column=5).value
            if cell_value != store_name:
                print(f"Missing store name at row {row_idx + i}, writing '{store_name}'")
                itemlist_sheet.cell(row=row_idx + i, column=5).value = store_name

    print("Store names in column E are ensured.")

# 바코드가 정렬된 후 데이터 입력
def update_itemlist(scraped_data, itemlist_sheet):
    store_names = ['Booragoon', 'Carousel', 'Northbridge', 'Innaloo']
    print(f"itemlist_sheet.max_row: {itemlist_sheet.max_row}")
    
    for row_idx in range(2, itemlist_sheet.max_row + 1, 4):
        barcode = str(itemlist_sheet.cell(row=row_idx, column=6).value)  # Booragoon 행의 바코드 확인 (column=6)
        print(f"Barcode in row {row_idx}: {barcode}")
        
        if barcode in scraped_data:
            product_data = scraped_data[barcode]
            for i in range(4):
                current_row = row_idx + i
                location = store_names[i]

                # 1. Title (B열) 입력 (Booragoon 행에만)
                if i == 0 and itemlist_sheet.cell(row=current_row, column=2).value is None:
                    itemlist_sheet.cell(row=current_row, column=2).value = product_data['title']

                # 2. A열이 비어있는 경우 B열의 값으로 채움
                if itemlist_sheet.cell(row=current_row, column=1).value is None:
                    itemlist_sheet.cell(row=current_row, column=1).value = itemlist_sheet.cell(row=current_row, column=2).value

                # 3. D열에 'Location' 문자열 입력 (Booragoon 행에만)
                if i == 0 and itemlist_sheet.cell(row=current_row, column=4).value is None:
                    print(f"Writing 'Location' to D{current_row}")
                    itemlist_sheet.cell(row=current_row, column=4).value = 'Location'

                # 4. G열에 바코드 + _BR, _CA, _NB, _IN 형식으로 입력
                if itemlist_sheet.cell(row=current_row, column=7).value is None:
                    itemlist_sheet.cell(row=current_row, column=7).value = f"{barcode}_{location[:2].upper()}"

                # 5. H열에 각 매장의 재고 정보 입력
                stock_column = f"{location[:2].lower()}_stock"
                if itemlist_sheet.cell(row=current_row, column=8).value is None:
                    itemlist_sheet.cell(row=current_row, column=8).value = product_data.get(stock_column, 0)

                # 6. I열에 각 매장의 할인 정보 입력 (None인 경우 빈 문자열로 처리)
                discount_column = f"{location[:2].lower()}_discount"
                if itemlist_sheet.cell(row=current_row, column=9).value is None:
                    itemlist_sheet.cell(row=current_row, column=9).value = product_data.get(discount_column, '')

                # 7. M, N, O, P열에 각각 'shopify', 'deny', 'manual', 'TRUE' 입력 (비어 있을 경우에만)
                if itemlist_sheet.cell(row=current_row, column=13).value is None:
                    itemlist_sheet.cell(row=current_row, column=13).value = 'shopify'
                if itemlist_sheet.cell(row=current_row, column=14).value is None:
                    itemlist_sheet.cell(row=current_row, column=14).value = 'deny'
                if itemlist_sheet.cell(row=current_row, column=15).value is None:
                    itemlist_sheet.cell(row=current_row, column=15).value = 'manual'
                if itemlist_sheet.cell(row=current_row, column=16).value is None:
                    itemlist_sheet.cell(row=current_row, column=16).value = 'TRUE'

                # 8. Q열에는 Booragoon 행에만 'TRUE' 입력
                if i == 0 and itemlist_sheet.cell(row=current_row, column=17).value is None:
                    itemlist_sheet.cell(row=current_row, column=17).value = 'TRUE'

                # 9. R열에 수식 입력 (비어 있을 경우에만)
                if itemlist_sheet.cell(row=current_row, column=18).value is None:
                    itemlist_sheet.cell(row=current_row, column=18).value = f'=IF(AND(S{current_row}="O", T{current_row}="O"), "active", "archived")'

                 # 10-13. V, W, X, Y열에 유통기한 정보 입력 (Booragoon 행에만, 비어 있을 경우에만)
                if i == 0:
                    if itemlist_sheet.cell(row=current_row, column=22).value is None:
                        itemlist_sheet.cell(row=current_row, column=22).value = (
                            f"Expiration date: {product_data['br_expiration']}" if product_data['br_expiration'] else ''
                        )
                    if itemlist_sheet.cell(row=current_row, column=23).value is None:
                        itemlist_sheet.cell(row=current_row, column=23).value = (
                            f"Expiration date: {product_data['ca_expiration']}" if product_data['ca_expiration'] else ''
                        )
                    if itemlist_sheet.cell(row=current_row, column=24).value is None:
                        itemlist_sheet.cell(row=current_row, column=24).value = (
                            f"Expiration date: {product_data['nb_expiration']}" if product_data['nb_expiration'] else ''
                        )
                    if itemlist_sheet.cell(row=current_row, column=25).value is None:
                        itemlist_sheet.cell(row=current_row, column=25).value = (
                            f"Expiration date: {product_data['in_expiration']}" if product_data['in_expiration'] else ''
                        )

    print("Itemlist has been successfully updated.")

# 실행 부분
check_duplicate_barcodes(itemlist_sheet)

itemlist_barcodes = get_itemlist_barcodes(itemlist_sheet)
barcodes_to_add, barcodes_to_delete = compare_barcodes(scraped_data, itemlist_barcodes)

# 바코드 삭제 및 추가
delete_barcodes(itemlist_barcodes, barcodes_to_delete, itemlist_sheet)
add_barcodes(barcodes_to_add, scraped_data, itemlist_sheet)

# 중복 바코드가 있는지 다시 확인
check_duplicate_barcodes(itemlist_sheet)

# 바코드가 추가되었는지 다시 확인
barcodes_to_add, barcodes_to_delete = compare_barcodes(scraped_data, get_itemlist_barcodes(itemlist_sheet))

if barcodes_to_add:
    print(f"Error: Some barcodes were not added properly: {barcodes_to_add}")
    raise Exception("Barcodes were not added correctly.")

# 바코드 재정렬
reorder_barcodes(scraped_data, itemlist_sheet)

# 매장 이름 확인 및 수정
ensure_store_names(itemlist_sheet)

# 데이터 입력 (Scraped 데이터를 기반으로 Itemlist 업데이트)
update_itemlist(scraped_data, itemlist_sheet)

# 엑셀 파일 저장
updated_itemlist_file = f'Updated_Itemlist_{current_date}.xlsx'
try:
    itemlist_wb.save(updated_itemlist_file)
    print(f"Updated Itemlist saved as '{updated_itemlist_file}'.")
except Exception as e:
    print(f"Error saving updated itemlist file: {e}")
