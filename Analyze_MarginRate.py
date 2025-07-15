import pandas as pd
import numpy as np
import datetime
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList

def calculate_margin_rate(itemlist_filename, marginrate_filename):
    # ----------------------------------------------------------------------
    # 1) 엑셀에서 필요한 열 읽어오기 (B=상품명, K=할인가, L=판매가, O=도매가)
    #    열 이름이 달라도 usecols="B,K,L,O" 또는 usecols=[1,10,11,14] 등으로 불러올 수 있음
    # ----------------------------------------------------------------------
    df = pd.read_excel(
        itemlist_filename, 
        header=0, 
        usecols="B,K,L,O"  # 열 위치에 따라 수정 가능
    )
    # 불러온 열에 새 이름 부여 (4개 열이라고 가정)
    df.columns = ["상품명", "할인가격", "판매가격", "도매가격"]

    # ----------------------------------------------------------------------
    # 2) 마진율 계산
    #    - 판매가 또는 도매가가 0이면 제외
    #    - 마진율이 0~500% 범위만 유효
    #    - 이후 마진율은 소수점 둘째 자리까지만 표기
    # ----------------------------------------------------------------------
    # (2-1) 0이 아닌 판매가/도매가만 필터
    df_valid = df[(df["판매가격"] != 0) & (df["도매가격"] != 0)].copy()

    # (2-2) 마진율 계산
    df_valid["마진율"] = ((df_valid["판매가격"] - df_valid["도매가격"]) 
                       / df_valid["판매가격"]) * 100

    # (2-3) 0% 이상 500% 이하만 유효
    condition = (df_valid["마진율"] >= 0) & (df_valid["마진율"] <= 500)
    df_included = df_valid[condition].copy()

    # (2-4) **마진율을 소수점 둘째 자리까지 반올림**  
    df_included["마진율"] = df_included["마진율"].round(2)

    # (2-5) 제외된 항목들 (판매/도매가 0인 항목 + 마진율 범위를 벗어난 항목)
    df_excluded = pd.concat([
        df[~df.index.isin(df_valid.index)],  # 판매/도매가 0
        df_valid[~condition]                 # 마진율 범위 벗어남
    ])

    # ----------------------------------------------------------------------
    # 3) MarginRate 시트용 / Exception 시트용
    # ----------------------------------------------------------------------
    df_marginrate = df_included[["상품명", "할인가격", "판매가격", "도매가격", "마진율"]]

    df_excluded = df_excluded.copy()
    if "마진율" not in df_excluded.columns:
        df_excluded["마진율"] = np.nan
    df_exception = df_excluded[["상품명", "할인가격", "판매가격", "도매가격", "마진율"]]

    # ----------------------------------------------------------------------
    # 4) Ranking 시트 (마진율 상위 50, 하위 50)
    # ----------------------------------------------------------------------
    df_sorted = df_marginrate.sort_values(by="마진율", ascending=False)
    df_top50 = df_sorted.head(50)
    df_bottom50 = df_sorted.tail(50).sort_values(by="마진율", ascending=True)

    # ----------------------------------------------------------------------
    # 5) Distribution 시트 (0~95~100% 구간까지만)
    #    - 마진율이 100%를 초과하는 경우는 제외
    #    - 구간: 0~5%, 5~10%, ... 95~100%
    # ----------------------------------------------------------------------
    df_dist = df_included[df_included["마진율"] <= 100].copy()  # 100% 이하만
    bins = list(range(0, 101, 5))  # 0,5,10,...,100
    labels = [f"{bins[i]}~{bins[i+1]}%" for i in range(len(bins)-1)]
    # 예: ["0~5%", "5~10%", ..., "95~100%"]

    # pd.cut으로 분포 구간 나누기 (right=True 이면 5% 구간은 (0,5], 100도 마지막 구간에 포함)
    df_dist["마진율 구간"] = pd.cut(
        df_dist["마진율"],
        bins=bins,
        labels=labels,
        include_lowest=True,  # 0 포함
        right=True            # 마지막 구간 (95,100] 에 100 포함
    )

    dist_count = df_dist.groupby("마진율 구간")["상품명"].count().reset_index()
    dist_count.columns = ["마진율 구간", "상품수"]

    # 구간(라벨)이 누락되지 않도록 하는 처리 (만약 빈 구간이 있을 경우 0개로 표기)
    full_dist_df = pd.DataFrame({"마진율 구간": labels})
    full_dist_df = pd.merge(full_dist_df, dist_count, on="마진율 구간", how="left")
    full_dist_df["상품수"] = full_dist_df["상품수"].fillna(0).astype(int)

    # ----------------------------------------------------------------------
    # 6) 결과 엑셀 파일 작성
    # ----------------------------------------------------------------------
    with pd.ExcelWriter(marginrate_filename, engine="xlsxwriter") as writer:
        # 1) MarginRate 시트
        df_marginrate.to_excel(writer, sheet_name="MarginRate", index=False)
        # 2) Exception 시트
        df_exception.to_excel(writer, sheet_name="Exception", index=False)
        # 3) Ranking 시트
        df_top50.to_excel(writer, sheet_name="Ranking", index=False, startrow=0)
        df_bottom50.to_excel(writer, sheet_name="Ranking", index=False, startrow=len(df_top50)+2)
        # 4) Distribution 시트
        full_dist_df.to_excel(writer, sheet_name="Distribution", index=False)

    # ----------------------------------------------------------------------
    # 7) 그래프 생성 (openpyxl)
    # ----------------------------------------------------------------------
    wb = load_workbook(marginrate_filename)
    ws_dist = wb["Distribution"]

    max_row = ws_dist.max_row
    chart = BarChart()
    chart.type = "col"
    chart.title = "마진율 구간별 상품수"
    chart.x_axis.title = "마진율 구간"
    chart.y_axis.title = "상품 수"

    # 카테고리(구간: A열)와 데이터(상품수: B열) 범위
    cats = Reference(ws_dist, min_col=1, min_row=2, max_row=max_row)  # A2:A??
    data = Reference(ws_dist, min_col=2, min_row=1, max_row=max_row)  # B1:B??
    chart.add_data(data, titles_from_data=True, from_rows=False)
    chart.set_categories(cats)
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True

    # 그래프를 시트 내 특정 위치(E2)에 추가
    ws_dist.add_chart(chart, "E2")
    wb.save(marginrate_filename)

if __name__ == "__main__":
    # 날짜 문자열 (yyMMdd), 예: 2025년 1월 14일 -> "250114"
    today_str = datetime.datetime.now().strftime('%y%m%d')
    itemlist_file = f"Itemlist_{today_str}.xlsx"
    marginrate_file = f"MarginRate_{today_str}.xlsx"

    calculate_margin_rate(itemlist_file, marginrate_file)
    print(f"마진율 계산 완료! '{marginrate_file}' 파일이 생성되었습니다.")
