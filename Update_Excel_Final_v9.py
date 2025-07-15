import openpyxl
from openpyxl import Workbook
import csv
import re
import os
import pandas as pd
from datetime import datetime, timedelta


# v9. 업데이트 내용
# Purchase 파일 내용이 업데이트되지 않던 문제 수정


current_date = datetime.now().strftime('%y%m%d')

scraped_file = f'Scraped_{current_date}.xlsx'
itemlist_file = f'Itemlist_{current_date}.xlsx'
scraped_m_file = f'ScrapedM_{current_date}.xlsx'
purchase_file = f'Purchase_{current_date}.xlsx'  # 추가된 Purchase 파일

print(f"Loading scraped file: {scraped_file}")


# ---------------------------------------------------------------------------
# 1) 파일 로딩
# ---------------------------------------------------------------------------
try:
    scraped_wb = openpyxl.load_workbook(scraped_file, read_only=False)
    scraped_sheet = scraped_wb.active
    print("Scraped file loaded successfully.")
except Exception as e:
    print(f"Error loading scraped file: {e}")
    scraped_sheet = None

print(f"Loading scrapedM file: {scraped_m_file}")
try:
    scraped_m_wb = openpyxl.load_workbook(scraped_m_file, read_only=False)
    scraped_m_sheet = scraped_m_wb.active
    print("ScrapedM file loaded successfully.")
except Exception as e:
    print(f"Error loading scrapedM file: {e}")
    scraped_m_sheet = None

print(f"Loading itemlist file: {itemlist_file}")
try:
    itemlist_wb = openpyxl.load_workbook(itemlist_file, read_only=False)
    itemlist_sheet = itemlist_wb.active
    print("Itemlist file loaded successfully.")
except Exception as e:
    print(f"Error loading itemlist file: {e}")
    itemlist_sheet = None

    print(f"Loading purchase file: {purchase_file}")
try:
    purchase_wb = openpyxl.load_workbook(purchase_file, read_only=False)
    purchase_sheet = purchase_wb.active
    print("Purchase file loaded successfully.")
except Exception as e:
    print(f"Error loading purchase file: {e}")
    purchase_sheet = None


# ---------------------------------------------------------------------------
# 2) scraped_data 구성
# ---------------------------------------------------------------------------
scraped_data = {}
if scraped_sheet:
    print("Reading scraped data...")
    try:
        for row in scraped_sheet.iter_rows(min_row=2, values_only=True):
            barcode = str(row[0])
            title = row[1]

            # 각 매장별 재고/할인/유통기한
            cr_stock = row[2] or 0
            cr_discount = row[3]
            cr_expiration = row[4]

            nb_stock = row[5] or 0
            nb_discount = row[6]
            nb_expiration = row[7]

            in_stock = row[8] or 0
            in_discount = row[9]
            in_expiration = row[10]

            br_stock = row[11] or 0
            br_discount = row[12]
            br_expiration = row[13]

            # 총 재고 계산
            total_stock = cr_stock + nb_stock + in_stock + br_stock

            # 할인 정보 및 유통기한 결정 (가장 많은 재고를 가진 할인 종료일 기준)
            expiration_infos = [
                (cr_expiration, cr_stock, cr_discount),
                (nb_expiration, nb_stock, nb_discount),
                (in_expiration, in_stock, in_discount),
                (br_expiration, br_stock, br_discount)
            ]

            # 가장 많은 재고를 가진 유통기한을 기준으로 결정
            exp_date_counter = {}
            for exp, stock, _ in expiration_infos:
                if exp:
                    exp_date_counter[exp] = exp_date_counter.get(exp, 0) + stock

            # 다수 유통기한 중 재고 많은 날짜를 우선으로 선택
            if exp_date_counter:
                selected_exp = max(exp_date_counter.items(), key=lambda x: x[1])[0]
                expiration_date = selected_exp
                discount_price = next((d for e, s, d in expiration_infos if e == selected_exp and d), None)
            else:
                expiration_date = None
                discount_price = None

            # 🔥 핵심: 매장별 재고는 항상 scraped에서 가져온 값으로 고정
            store_stocks = {
                'CR': cr_stock,
                'NB': nb_stock,
                'IN': in_stock,
                'BR': br_stock
            }

            # scraped_data에 저장
            scraped_data[barcode] = {
                'title': title,
                'total_stock': total_stock,
                'discount_price': discount_price,
                'expiration_date': expiration_date,
                'store_stocks': store_stocks
            }

        print("Scraped data read successfully.")
    except Exception as e:
        print(f"Error reading scraped data: {e}")


def read_scraped_m_data(scraped_m_sheet):
    scraped_m_data = {}
    for row in scraped_m_sheet.iter_rows(min_row=2, values_only=True):
        barcode = str(row[1])
        wholesaler_price = row[2]
        retail_price = row[3]
        scraped_m_data[barcode] = {
            'wholesaler_price': wholesaler_price,
            'retail_price': retail_price
        }
    return scraped_m_data


# ---------------------------------------------------------------------------
# 3) 바코드 관련 함수
# ---------------------------------------------------------------------------
def get_itemlist_barcodes(itemlist_sheet):
    itemlist_barcodes = {}
    for row_idx in range(2, itemlist_sheet.max_row + 1):
        cell_val = itemlist_sheet.cell(row=row_idx, column=14).value
        if cell_val is None:
            continue
        # 숫자(float/int)인 경우 int 변환 -> str
        if isinstance(cell_val, (int, float)):
            cell_val = str(int(cell_val))  # 123456.0 -> 123456
        else:
            cell_val = str(cell_val).strip()  # 문자열
        barcode = cell_val
        itemlist_barcodes[barcode] = row_idx
    return itemlist_barcodes

def check_duplicate_barcodes(itemlist_sheet):
    barcodes_seen = set()
    duplicate_found = False
    for row_idx in range(2, itemlist_sheet.max_row + 1):
        barcode = str(itemlist_sheet.cell(row=row_idx, column=14).value)
        if barcode in barcodes_seen:
            print(f"Duplicate barcode found: {barcode} at row {row_idx}")
            duplicate_found = True
            exit(1)
        barcodes_seen.add(barcode)
    
    if not duplicate_found:
        print("중복된 바코드가 없습니다.")

def compare_barcodes(scraped_data, itemlist_barcodes):
    scraped_barcodes = set(scraped_data.keys())
    itemlist_barcodes_set = set(itemlist_barcodes.keys())
    
    barcodes_to_add = scraped_barcodes - itemlist_barcodes_set
    barcodes_to_delete = itemlist_barcodes_set - scraped_barcodes
    return barcodes_to_add, barcodes_to_delete


# ---------------------------------------------------------------------------
# 4) add_barcodes / reorder_barcodes / archive_barcodes (최종)
# ---------------------------------------------------------------------------
def add_barcodes(barcodes_to_add, scraped_data, itemlist_sheet):
    for barcode in barcodes_to_add:
        new_row = [
            scraped_data[barcode]['title'],
            scraped_data[barcode]['title'],
            None,
            'TRUE',
            None,
            None,
            'shopify',
            scraped_data[barcode]['total_stock'],
            'deny',
            'manual',
            None,
            None,
            'TRUE',
            str(barcode),
            None,
            None,
            None,
            None,
            None,
            None
        ]
        last_row = itemlist_sheet.max_row + 1
        print(f"Adding new barcode {barcode} at row {last_row}")
        for col_idx, value in enumerate(new_row, start=1):
            cell = itemlist_sheet.cell(row=last_row, column=col_idx, value=value)
            if col_idx in [11, 12, 15]:
                if value is not None:
                    cell.number_format = '0.00'

def reorder_barcodes(scraped_data, sheet):
    real_max_col = sheet.max_column
    forced_min_col = 25  #강제로 25열까지 선택해서 이동
    max_col = max(real_max_col, forced_min_col)

    item_barcodes = get_itemlist_barcodes(sheet)
    new_order = []
    for bc in scraped_data.keys():
        if bc in item_barcodes:
            row_idx = item_barcodes[bc]
            row_data = []
            for c in range(1, max_col + 1):
                val = sheet.cell(row=row_idx, column=c).value
                row_data.append(val)
            new_order.append(row_data)

    sheet.delete_rows(2, sheet.max_row - 1)
    for row_data in new_order:
        sheet.append(row_data)

def archive_barcodes_first(barcodes_to_delete, sheet):
    """
    (1) 바코드 to_delete 행을 전부 rows_to_archive 에 보관하고,
    (2) 시트에서 제거한 뒤,
    (3) rows_to_archive를 리턴한다. (아직 archived처리는 안 함)
    """
    real_max_col = sheet.max_column
    forced_min_col = 25
    max_col = max(real_max_col, forced_min_col)

    item_barcodes = get_itemlist_barcodes(sheet)
    rows_to_archive = []
    for bc in barcodes_to_delete:
        if bc in item_barcodes:
            row_idx = item_barcodes[bc]
            row_data = []
            for c in range(1, max_col + 1):
                row_data.append(sheet.cell(row=row_idx, column=c).value)
            rows_to_archive.append((bc, row_data))

    # 행 삭제
    rows_to_archive.sort(key=lambda x: item_barcodes[x[0]], reverse=True)
    for bc, _ in rows_to_archive:
        row_idx = item_barcodes[bc]
        sheet.delete_rows(row_idx)
    return rows_to_archive

def insert_archived_rows(rows_to_archive, sheet):
    """
    rows_to_archive에 보관된 행을 시트 맨 아래로 삽입,
    P열(16)=archived, C열(3)=...,deleted
    """
    last_row = sheet.max_row
    for bc, row_data in rows_to_archive:
        last_row += 1
        for c, val in enumerate(row_data, start=1):
            sheet.cell(row=last_row, column=c, value=val)
        
        # archived 처리
        sheet.cell(row=last_row, column=16).value = "archived"

        # Q/R 열 (17, 18열)에 O가 아닌 값 넣기 → active로 분류되지 않도록
        sheet.cell(row=last_row, column=17).value = "-"  # Q열
        sheet.cell(row=last_row, column=18).value = "-"  # R열

        tags_cell = sheet.cell(row=last_row, column=3)
        if tags_cell.value:
            if "deleted" not in str(tags_cell.value).lower():
                tags_cell.value = str(tags_cell.value) + ", deleted"
        else:
            tags_cell.value = "deleted"


# ---------------------------------------------------------------------------
# 5) 기존 행 업데이트 / update_prices / discount ...
# ---------------------------------------------------------------------------
def update_prices(scraped_m_data, itemlist_sheet):
    itemlist_barcodes = get_itemlist_barcodes(itemlist_sheet)
    for barcode, row_idx in itemlist_barcodes.items():
        if barcode in scraped_m_data:
            data = scraped_m_data[barcode]
            k_cell = itemlist_sheet.cell(row=row_idx, column=11)
            l_cell = itemlist_sheet.cell(row=row_idx, column=12)
            o_cell = itemlist_sheet.cell(row=row_idx, column=15)

            k_cell.value = data['retail_price']
            l_cell.value = data['retail_price']
            k_cell.number_format = '0.00'
            l_cell.number_format = '0.00'

            # wholesaler_price가 0이 아닐 때만 O열에 값 작성
            if data['wholesaler_price'] != 0:
                o_cell.value = data['wholesaler_price']
                o_cell.number_format = '0.00'
            # 0이면 O열에 아무것도 하지 않음 (기존 값 유지)
        else:
            # ScrapedM에 해당 바코드가 없을 경우
            itemlist_sheet.cell(row=row_idx, column=11).value = 0
            itemlist_sheet.cell(row=row_idx, column=12).value = 0
            itemlist_sheet.cell(row=row_idx, column=15).value = 0
            print(f"Barcode {barcode} not found in ScrapedM. Prices set to 0.")

def update_existing_rows(scraped_data, itemlist_sheet):
    print("Matching and updating itemlist data...")
    try:
        for row_idx in range(2, itemlist_sheet.max_row + 1):
            variant_barcode = str(itemlist_sheet.cell(row=row_idx, column=14).value)
            if variant_barcode in scraped_data:
                row = itemlist_sheet[row_idx]
                if row[1].value != scraped_data[variant_barcode]['title']:
                    row[1].value = scraped_data[variant_barcode]['title']
                row[7].value = scraped_data[variant_barcode]['total_stock']
    except Exception as e:
        print(f"Error updating itemlist data: {e}")

def update_discount(scraped_data, itemlist_barcodes, itemlist_sheet):
    for barcode, data in scraped_data.items():
        if barcode in itemlist_barcodes:
            row_idx = itemlist_barcodes[barcode]
            if 'discount_price' in data and data['discount_price'] is not None:
                itemlist_sheet.cell(row=row_idx, column=11).value = data['discount_price']

from datetime import datetime, timedelta

from datetime import datetime, timedelta

def update_expiration(scraped_data, itemlist_barcodes, itemlist_sheet):
    today = datetime.today()
    six_months_later = today + timedelta(days=183)  # 약 6개월 후

    for barcode, data in scraped_data.items():
        if barcode in itemlist_barcodes:
            row_idx = itemlist_barcodes[barcode]
            expiration_date = data.get('expiration_date')
            discount_price = data.get('discount_price')

            if not discount_price:
                # 할인 정보가 없으면 무조건 '-'
                itemlist_sheet.cell(row=row_idx, column=20).value = "-"
                continue

            if expiration_date:
                try:
                    exp_dt = datetime.strptime(expiration_date, "%d/%m/%Y")
                    if exp_dt < today or exp_dt > six_months_later:
                        itemlist_sheet.cell(row=row_idx, column=20).value = "-"
                    else:
                        expiration_text = f"Expiration date: {expiration_date}"
                        itemlist_sheet.cell(row=row_idx, column=20).value = expiration_text
                except ValueError:
                    itemlist_sheet.cell(row=row_idx, column=20).value = "-"
            else:
                itemlist_sheet.cell(row=row_idx, column=20).value = "-"

def update_store_stock_status(scraped_data, itemlist_barcodes, itemlist_sheet):
    headers = {
        21: 'Stock Status CR (product.metafields.custom.stock_cr)',
        22: 'Stock Status NB (product.metafields.custom.stock_nb)',
        23: 'Stock Status IN (product.metafields.custom.stock_in)',
        24: 'Stock Status BR (product.metafields.custom.stock_br)'
    }
    for col_idx, header in headers.items():
        itemlist_sheet.cell(row=1, column=col_idx).value = header

    for barcode, data in scraped_data.items():
        if barcode in itemlist_barcodes:
            row_idx = itemlist_barcodes[barcode]
            store_stocks = data['store_stocks']
            store_codes = ['CR', 'NB', 'IN', 'BR']
            for idx, store_code in enumerate(store_codes, start=21):
                stock = store_stocks[store_code]
                if stock == 0:
                    status = 'Out of stock'
                elif stock >= 10:
                    status = 'In stock'
                else:
                    status = f'{stock} Left'
                itemlist_sheet.cell(row=row_idx, column=idx).value = status

def update_variant_sku(itemlist_sheet):
    for row_idx in range(2, itemlist_sheet.max_row + 1):
        value = itemlist_sheet.cell(row=row_idx, column=14).value  # N열의 값 읽기
        itemlist_sheet.cell(row=row_idx, column=5).value = value    # E열에 값 직접 할당

def update_status_column(itemlist_sheet):
    print("Updating Status column (P열)...")
    try:
        for row_idx in range(2, itemlist_sheet.max_row + 1):
            status_formula = f'=IF(AND(Q{row_idx}="O", R{row_idx}="O"), "active", "archived")'
            itemlist_sheet.cell(row=row_idx, column=16).value = status_formula
    except Exception as e:
        print(f"Error updating Status column: {e}")

def adjust_prices_for_zero_stock(itemlist_sheet):
    print("Adjusting prices for items with zero stock...")
    try:
        for row_idx in range(2, itemlist_sheet.max_row + 1):
            inventory_qty = itemlist_sheet.cell(row=row_idx, column=8).value
            if inventory_qty == 0 or inventory_qty == '0':
                variant_compare_at_price = itemlist_sheet.cell(row=row_idx, column=12).value
                itemlist_sheet.cell(row=row_idx, column=11).value = variant_compare_at_price
    except Exception as e:
        print(f"Error adjusting prices: {e}")

def read_purchase_data(purchase_sheet):
    """
    Purchase 시트에서 description을 정규화한 후, dict로 저장
    """
    purchase_data = {}  
    for row in purchase_sheet.iter_rows(min_row=2, values_only=True):
        description = str(row[0]) if row[0] else ""
        normalized_desc = re.sub(r"\(.*?\)|\[.*?\]", "", description).strip()
        price = row[2]
        comment = row[4] if len(row) > 4 else None
        purchase_data[normalized_desc] = {
            'price': price,
            'comment': comment
        }
    return purchase_data

def update_cost_per_item_from_purchase(purchase_data, itemlist_sheet):
    """
    Itemlist의 Title(B열)을 정규화하여 Purchase 데이터와 매칭
    매칭되지 않은 항목은 누락 리스트에 기록
    """
    unmatched_rows = []

    for row_idx in range(2, itemlist_sheet.max_row + 1):
        original_title = itemlist_sheet.cell(row=row_idx, column=2).value  # B열 Title
        if original_title is None:
            continue
        
        # 정규화
        normalized_title = re.sub(r"\(.*?\)|\[.*?\]", "", str(original_title)).strip()
        
        if normalized_title in purchase_data:
            cost_cell = itemlist_sheet.cell(row=row_idx, column=15)  # O열: Cost per item
            wholesaler_cell = itemlist_sheet.cell(row=row_idx, column=25)  # Y열

            cost_value = purchase_data[normalized_title]['price']
            comment_value = purchase_data[normalized_title]['comment']

            cost_cell.value = cost_value
            cost_cell.number_format = '0.00'

            wholesaler_cell.value = comment_value if comment_value else None
        else:
            # 매칭 실패 시 기록
            unmatched_rows.append({
                "Row": row_idx,
                "Title": original_title,
                "Normalized Title": normalized_title
            })
    
    # 누락 항목을 Excel로 저장
    if unmatched_rows:
        df_unmatched = pd.DataFrame(unmatched_rows)
        df_unmatched.to_excel("Unmatched_Purchase_Titles.xlsx", index=False)
        print(f"[알림] 매칭되지 않은 {len(unmatched_rows)}개의 항목을 'Unmatched_Purchase_Titles.xlsx'로 저장했습니다.")
    else:
        print("[완료] 모든 항목이 Purchase 정보와 매칭되었습니다.")

def update_special_tag(itemlist_sheet):
    """
    1) Variant Price(K열, col=11)와 Variant Compare At Price(L열, col=12)를 비교
       - 같다면 C열에 ' Special'이라는 텍스트가 없어야 함(이미 있으면 제거)
       - 다르면 C열에 ' Special'이 있어야 함(없으면 추가)
    """
    for row_idx in range(2, itemlist_sheet.max_row + 1):
        # 각각의 셀 값 가져오기
        variant_price = itemlist_sheet.cell(row=row_idx, column=11).value
        compare_price = itemlist_sheet.cell(row=row_idx, column=12).value
        tags_cell = itemlist_sheet.cell(row=row_idx, column=3)

        # Tags 셀이 None인 경우 빈 문자열로 처리
        if tags_cell.value is None:
            tags_cell.value = ""

        # 문자열 형태로 변환
        tags_str = str(tags_cell.value)
        special_text = ", Special"

        # (1) 가격이 같으면 ', Special' 제거
        if variant_price == compare_price:
            if special_text in tags_str:
                new_tags_str = tags_str.replace(special_text, "")
                new_tags_str = new_tags_str.strip()  # 앞뒤 공백 정리
                tags_cell.value = new_tags_str
        else:
            # (2) 가격이 다르면 ', Special' 추가
            if special_text not in tags_str:
                # 기존 태그가 비어있지 않다면, 앞에 공백 포함해서 붙일지 여부 결정
                # 여기서는 바로 이어붙이는 방식으로 예시
                new_tags_str = tags_str.strip() + special_text
                tags_cell.value = new_tags_str


# ---------------------------------------------------------------------------
# 6) CSV 파일 생성 관련 함수
# ---------------------------------------------------------------------------
def get_active_rows_in_memory(itemlist_sheet):
    """
    Itemlist 시트에서 Q,R열 모두 "O"인 행들을 'active' 상태로 간주하고,
    (headers, active_rows)를 리턴한다.
    실제 CSV 파일 작성은 하지 않는다.
    """
    print("Collecting 'active' rows in memory based on Q and R columns...")
    try:
        # 헤더
        headers = [cell.value for cell in itemlist_sheet[1]]
        active_rows = []

        # 본문 데이터 추출
        for row_idx, row in enumerate(itemlist_sheet.iter_rows(min_row=2, values_only=True), start=2):
            q_value = str(row[16]).strip() if row[16] is not None else ""
            r_value = str(row[17]).strip() if row[17] is not None else ""
            if q_value == "O" and r_value == "O":
                row = list(row)
                # 바코드(N열, 인덱스 13) 숫자 포맷
                if row[13] is not None:
                    row[13] = f'="{str(row[13])}"'
                # 숫자 포맷 (K=10, L=11, O=14 인덱스)
                for col_idx in [10, 11, 14]:
                    if row[col_idx] is not None and isinstance(row[col_idx], (float, int)):
                        row[col_idx] = f"{row[col_idx]:.2f}"
                # P열(인덱스 15)에 'active'
                row[15] = "active"
                active_rows.append(row)

        print(f"Active rows collected in memory. Count: {len(active_rows)}")
        return headers, active_rows

    except Exception as e:
        print(f"Error collecting active rows in memory: {e}")
        return [], []

def get_archived_rows_in_memory(itemlist_sheet):
    """
    Itemlist 시트에서 Q열 또는 R열 중 하나라도 'O'가 아닌 행을 'archived'로 간주하고,
    (headers, archived_rows)를 리턴한다.
    실제 CSV 파일 작성은 하지 않는다.
    """
    print("Collecting 'archived' rows in memory based on Q and R columns...")
    try:
        headers = [cell.value for cell in itemlist_sheet[1]]
        archived_rows = []

        for row_idx, row in enumerate(itemlist_sheet.iter_rows(min_row=2, values_only=True), start=2):
            q_value = str(row[16]).strip() if row[16] is not None else ""
            r_value = str(row[17]).strip() if row[17] is not None else ""
            if q_value != "O" or r_value != "O":
                row = list(row)
                # 바코드(N열, 인덱스 13) 숫자 포맷
                if row[13] is not None:
                    row[13] = f'="{str(row[13])}"'
                # 숫자 포맷 (K=10, L=11, O=14 인덱스)
                for col_idx in [10, 11, 14]:
                    if row[col_idx] is not None and isinstance(row[col_idx], (float, int)):
                        row[col_idx] = f"{row[col_idx]:.2f}"
                # P열(인덱스 15)에 'archived'
                row[15] = "archived"
                archived_rows.append(row)

        print(f"Archived rows collected in memory. Count: {len(archived_rows)}")
        return headers, archived_rows

    except Exception as e:
        print(f"Error collecting archived rows in memory: {e}")
        return [], []

def save_active_csv(headers, active_rows, output_csv_file):
    """
    이미 메모리에 수집된 active_rows를 CSV로 저장한다.
    """
    print(f"Saving {len(active_rows)} active rows to '{output_csv_file}'...")
    try:
        with open(output_csv_file, mode='w', newline='', encoding='utf-8-sig') as csvfile:
            writer = csv.writer(csvfile, quoting=csv.QUOTE_MINIMAL)
            writer.writerow(headers)     # 헤더
            writer.writerows(active_rows)
        print("Active CSV saved successfully.")
    except Exception as e:
        print(f"Error saving active CSV: {e}")

def save_total_csv(headers, active_rows, archived_rows, output_csv_file):
    """
    Active + Archived 데이터를 합쳐서 Total CSV 파일로 생성한다.
    """
    print(f"Saving Total CSV to '{output_csv_file}' (Active + Archived)...")

    try:
        with open(output_csv_file, mode='w', newline='', encoding='utf-8-sig') as csvfile:
            writer = csv.writer(csvfile, quoting=csv.QUOTE_MINIMAL)

            # 헤더
            writer.writerow(headers)

            # Active 행
            for row in active_rows:
                writer.writerow(row)

            # Archived 행
            for row in archived_rows:
                writer.writerow(row)

        print(f"Total CSV saved successfully. (Rows: {len(active_rows) + len(archived_rows)})")

    except Exception as e:
        print(f"Error saving total CSV: {e}")


# ---------------------------------------------------------------------------
# 메인 실행 흐름
# ---------------------------------------------------------------------------
if itemlist_sheet:
    # (1) 중복 바코드 검사
    check_duplicate_barcodes(itemlist_sheet)

    # (2) 바코드 비교
    item_barcodes = get_itemlist_barcodes(itemlist_sheet)
    to_add, to_delete = compare_barcodes(scraped_data, item_barcodes)

    # (3) 삭제 대상 바코드 먼저 rows_to_archive에 저장하고 시트에서 삭제
    rows_to_archive = archive_barcodes_first(to_delete, itemlist_sheet)

    # (4) 새 바코드 추가
    add_barcodes(to_add, scraped_data, itemlist_sheet)

    # (5) 바코드 재정렬
    reorder_barcodes(scraped_data, itemlist_sheet)
    item_barcodes = get_itemlist_barcodes(itemlist_sheet)

    # (6) 기존 행 업데이트
    update_existing_rows(scraped_data, itemlist_sheet)

    # (7) 가격/discount/expiration/storeStock/purchase ... 업데이트
    scraped_m_data = read_scraped_m_data(scraped_m_sheet)
    update_prices(scraped_m_data, itemlist_sheet)
    item_barcodes = get_itemlist_barcodes(itemlist_sheet)
    update_discount(scraped_data, get_itemlist_barcodes(itemlist_sheet), itemlist_sheet)
    
    update_expiration(scraped_data, item_barcodes, itemlist_sheet)
    update_store_stock_status(scraped_data, item_barcodes, itemlist_sheet)
    update_status_column(itemlist_sheet)
    update_variant_sku(itemlist_sheet)
    adjust_prices_for_zero_stock(itemlist_sheet)
    update_special_tag(itemlist_sheet)

    # Purchase 정보 업데이트
    if purchase_sheet:
        purchase_data = read_purchase_data(purchase_sheet)
        update_cost_per_item_from_purchase(purchase_data, itemlist_sheet)

    # (8) 삭제된 행 추가 및 archive 처리
    insert_archived_rows(rows_to_archive, itemlist_sheet)

    # (8-1) 다시 Status 업데이트 (방금 삽입된 archived 행 포함)
    update_status_column(itemlist_sheet)

    # (9) 최종 저장 + CSV
    temp_file = itemlist_file.replace(".xlsx","_temp.xlsx") 
    itemlist_wb.save(temp_file)
    if os.path.exists(itemlist_file):
        os.remove(itemlist_file)
    os.rename(temp_file, itemlist_file)

    headers_active, active_rows = get_active_rows_in_memory(itemlist_sheet)
    headers_archived, archived_rows = get_archived_rows_in_memory(itemlist_sheet)
    headers = headers_active

    active_csv_file = f'Active_Items_{current_date}.csv'
    save_active_csv(headers, active_rows, active_csv_file)

    total_csv_file = f'Total_Items_{current_date}.csv'
    save_total_csv(headers, active_rows, archived_rows, total_csv_file)